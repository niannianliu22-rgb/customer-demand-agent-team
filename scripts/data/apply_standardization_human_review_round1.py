#!/usr/bin/env python3
"""Apply only the explicit Standardization Human Review Round 1 decisions."""
from __future__ import annotations

import csv
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[2]
RUN_ID='RUN-202608-DEMAND-001'
ART=ROOT/f'runs/{RUN_ID}/artifacts'
DATA=ART/'unified_dataset.csv'; XLSX=ART/'unified_dataset.xlsx'; OUT=ROOT/'quality'
SCHOOL=ROOT/'config/data/school_human_review_round1.json'
DATES=ROOT/'config/data/date_human_review_round1.json'
SCHEMA=ROOT/'schemas/canonical_schema.json'

def sha(p): return hashlib.sha256(p.read_bytes()).hexdigest()
def read(p):
    with p.open(encoding='utf-8-sig',newline='') as h: return list(csv.DictReader(h))
def write(p,fields,rows):
    with p.open('w',encoding='utf-8-sig',newline='') as h:
        w=csv.DictWriter(h,fieldnames=fields);w.writeheader();w.writerows(rows)

def main():
    before={str(p.relative_to(ROOT)):sha(p) for p in [DATA,XLSX,SCHOOL,DATES,SCHEMA]}
    school=json.loads(SCHOOL.read_text(encoding='utf-8'))
    dates=json.loads(DATES.read_text(encoding='utf-8'))
    rows=read(DATA); fields=list(rows[0])
    for f in ['school_rule_id','school_rule_version','school_standardization_status','ddl_components']:
        if f not in fields: fields.append(f)
    applied_school=[]; applied_dates=[]
    for row in rows:
        raw=row['school_original']
        if raw in school['approved_mappings']:
            previous=row['school']; row['school']=school['approved_mappings'][raw]
            row['school_rule_id']=school['rule_id']; row['school_rule_version']=school['version']; row['school_standardization_status']='STANDARDIZED'
            applied_school.append({'record_id':f"{row['source_id']}:{row['source_row_id']}",'raw_school':raw,'previous_school':previous,'manual_decision':'MAP_TO_CANONICAL','final_school':row['school'],'classification':'CANONICAL_SCHOOL','rule_id':school['rule_id'],'rule_version':school['version']})
        elif raw in school['approved_non_school']:
            previous=row['school']; row['school']='NON_SCHOOL'
            row['school_rule_id']=school['rule_id']; row['school_rule_version']=school['version']; row['school_standardization_status']=school['approved_non_school'][raw]
            applied_school.append({'record_id':f"{row['source_id']}:{row['source_row_id']}",'raw_school':raw,'previous_school':previous,'manual_decision':'NON_SCHOOL','final_school':'NON_SCHOOL','classification':school['approved_non_school'][raw],'rule_id':school['rule_id'],'rule_version':school['version']})
        elif row['school']=='UNSTANDARDIZED':
            row['school_rule_id']='RULE-013'; row['school_rule_version']='1.1'; row['school_standardization_status']='REVIEW_REQUIRED'
        elif row['school']=='UNRESOLVED':
            row['school_rule_id']='RULE-013'; row['school_rule_version']='1.1'; row['school_standardization_status']='UNRESOLVED'
        elif row['school'] in {'NON_SCHOOL','NON_UNIVERSITY_ENTITY'}:
            row['school_rule_id']='RULE-013'; row['school_rule_version']='1.1'; row['school_standardization_status']=row['school']
        elif row['school']=='UNKNOWN' or not row['school']:
            row['school_rule_id']='RULE-013'; row['school_rule_version']='1.1'; row['school_standardization_status']='UNKNOWN'
        else:
            row['school_rule_id']='RULE-013'; row['school_rule_version']='1.1'; row['school_standardization_status']='STANDARDIZED'
        for field in ['consultation_date','ddl']:
            key=f"{row['source_id']}:{row['source_row_id']}:{field}"
            decision=dates['approved_decisions'].get(key)
            if not decision: continue
            previous=row[field]
            if decision['decision']=='CONFIRM_DATE':
                row[field]=decision['date']; row['ddl_components']=';'.join(decision.get('ddl_components',[])) if field=='ddl' else row.get('ddl_components','')
            applied_dates.append({'record_id':f"{row['source_id']}:{row['source_row_id']}",'field_name':field,'raw_value':row[f'{field}_original'],'previous_value':previous,'manual_decision':decision['decision'],'final_value':row[field],'ddl_components':row.get('ddl_components',''),'rule_id':'date_human_review_round1','rule_version':dates['version']})
    write(DATA,fields,rows)
    wb=load_workbook(XLSX); ws=wb.active
    old_headers=[c.value for c in ws[1]]
    for idx,name in enumerate(fields,1): ws.cell(1,idx).value=name
    for row_num,row in enumerate(rows,2):
        for col_num,name in enumerate(fields,1): ws.cell(row_num,col_num).value=row.get(name,'')
    # Clear only obsolete trailing header/value cells if the prior sheet had more columns.
    for col_num in range(len(fields)+1,len(old_headers)+1):
        for row_num in range(1,ws.max_row+1): ws.cell(row_num,col_num).value=None
    wb.save(XLSX)
    OUT.mkdir(exist_ok=True)
    write(OUT/'school_human_review_round1_applied.csv',['record_id','raw_school','previous_school','manual_decision','final_school','classification','rule_id','rule_version'],applied_school)
    write(OUT/'date_human_review_round1_applied.csv',['record_id','field_name','raw_value','previous_value','manual_decision','final_value','ddl_components','rule_id','rule_version'],applied_dates)
    audit={'run_id':RUN_ID,'artifact':'standardization_human_review_round1_application','applied_at':datetime.now(timezone.utc).isoformat(),'school':{'applied_rows':len(applied_school),'canonical_rows':sum(x['manual_decision']=='MAP_TO_CANONICAL' for x in applied_school),'non_school_rows':sum(x['manual_decision']=='NON_SCHOOL' for x in applied_school),'remaining_review_required_rows':sum(x['school']=='UNSTANDARDIZED' for x in rows)},'date':{'applied_decisions':len(applied_dates),'confirmed_dates':sum(x['manual_decision']=='CONFIRM_DATE' for x in applied_dates),'invalid_dates':sum(x['manual_decision']=='INVALID_DATE' for x in applied_dates),'multi_ddl_preserved':next(x['ddl_components'] for x in applied_dates if x['record_id']=='source_003:132')},'schema_version':'1.4.0','before_checksums':before,'after_checksums':{str(p.relative_to(ROOT)):sha(p) for p in [DATA,XLSX,SCHEMA]},'frozen_business_rules_modified':False,'not_modified':['Historical Demand conclusions','Academic Calendar','Demand Pattern','Opportunity results']}
    (OUT/'standardization_human_review_round1_application_audit.json').write_text(json.dumps(audit,ensure_ascii=False,indent=2),encoding='utf-8')
    print(json.dumps(audit['school']|audit['date'],ensure_ascii=False))
if __name__=='__main__': main()
