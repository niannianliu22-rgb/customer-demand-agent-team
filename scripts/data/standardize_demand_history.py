"""Deterministically standardize one customer-demand run without touching raw Excel.

This script consumes only run artifacts, the frozen canonical schema and the
ACTIVE business-rules registry.  It writes derived run artifacts; it never
writes under input/ or data/processed/.
"""

from __future__ import annotations

import csv
import datetime as dt
import hashlib
import json
import re
import sys
import os
from collections import Counter
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path

import yaml
from openpyxl import Workbook, load_workbook


ROOT = Path(__file__).resolve().parents[2]
RUN_ID = os.environ["CDAT_RUN_ID"]
ACTIVE_STATUSES = {"CONFIRMED", "EXCLUDED_BY_BUSINESS_RULE"}
DATE_DOT = re.compile(r"(?P<month>\d{1,2})\.(?P<day>\d{1,2})")
DATE_MONTH = re.compile(r"(?P<month>\d{1,2})月份?")
ISO_DATE = re.compile(r"^\d{4}-\d{1,2}-\d{1,2}$")
DEGREE_MAP = {"大一": "本科", "大二": "本科", "大三": "本科", "大四": "本科", "本科": "本科", "硕士": "硕士", "博士": "博士", "高中": "高中"}


def empty(value) -> bool:
    return value is None or (isinstance(value, str) and not value.strip())


def checksum(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def business_rules_identity(content: str) -> tuple[str, str]:
    """Use the frozen declared version, or its checksum when no version exists."""
    match = re.search(r"Business Rules Version:\s*([^*\n]+)", content)
    return (match.group(1).strip(), 'DECLARED_VERSION') if match else (checksum_value(content), 'CHECKSUM')


def checksum_value(content: str) -> str:
    return hashlib.sha256(content.encode('utf-8')).hexdigest()


def iso_date(value, year: int):
    """Return (standardized, transformation_kind, review_reason)."""
    if empty(value):
        return None, "blank", None
    if isinstance(value, (dt.datetime, dt.date)):
        return value.date().isoformat() if isinstance(value, dt.datetime) else value.isoformat(), "existing_date", None
    # Excel commonly stores an M.D expression such as 8.1 as a numeric cell.
    # Its displayed lexical form still fits RULE-002; values outside the M.D
    # shape (for example 0.615384...) continue to require human review.
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        text = str(value)
    elif isinstance(value, str):
        text = value.strip()
    else:
        return None, None, "date value is neither a date nor a rule-covered text pattern"
    if ISO_DATE.fullmatch(text):
        try:
            return dt.date.fromisoformat(text).isoformat(), "existing_iso_date", None
        except ValueError:
            return None, None, "invalid ISO date"
    match = DATE_DOT.search(text)
    if match:
        try:
            result = dt.date(year, int(match["month"]), int(match["day"])).isoformat()
            if "-" in text[match.end():] or "-" in text[:match.start()]:
                return result, "RULE-005_date_range_take_start", None
            if text != match.group(0):
                return result, "RULE-003_date_with_trailing_text", None
            return result, "RULE-002_month_dot_day", None
        except ValueError:
            return None, None, "invalid M.D date"
    match = DATE_MONTH.search(text)
    if match:
        try:
            return dt.date(year, int(match["month"]), 1).isoformat(), "RULE-004_month_only", None
        except ValueError:
            return None, None, "invalid month-only date"
    return None, None, "date text is not covered by RULE-002 through RULE-005"


def decimal_value(value):
    if empty(value):
        return None, None
    if isinstance(value, bool):
        return None, "boolean is not a valid amount"
    text = str(value).strip().replace(",", "").replace("¥", "").replace("￥", "")
    text = text.replace("AUD", "").replace("CNY", "").replace("RMB", "").strip()
    try:
        amount = Decimal(text)
    except InvalidOperation:
        return None, "amount is not a parseable numeric value"
    return amount, None


def row_issue(issues, source_id, source_row_id, field, original_value, reason, rule):
    issues.append({
        "issue_id": f"STD-REVIEW-{source_id}-{source_row_id}-{field}",
        "status": "STANDARDIZATION_REVIEW_REQUIRED",
        "source_id": source_id,
        "source_row_id": source_row_id,
        "field": field,
        "original_value": original_value.isoformat() if isinstance(original_value, (dt.date, dt.datetime)) else original_value,
        "reason": reason,
        "applicable_rule": rule,
        "recommended_action": "Human confirmation required; do not infer a new standardization rule."
    })


def load_json(path: Path):
    return json.loads(path.read_text(encoding="utf-8"))


def assert_snapshot_rule(manifest: dict, rule_name: str, path: Path) -> None:
    """Refuse to execute when a live frozen rule diverges from this Run."""
    entry = next(item for item in manifest["rule_versions"] if item["rule_name"] == rule_name)
    if checksum(path) != entry["checksum"]:
        raise RuntimeError(f"{rule_name} checksum diverges from the frozen Run snapshot")


def load_dimension_rules(run_manifest: dict) -> dict:
    """Load only human-confirmed, frozen dimension mappings for A05."""
    paths = {
        "school": ROOT / "config/data/school_aliases.yaml",
        "task": ROOT / "config/dimensions/task_type/task_type_rules_frozen_v17.yaml",
        "channel": ROOT / "config/dimensions/channel/channel_rules_frozen_v1.yaml",
        "date": ROOT / "config/data/date_standardization_rules_v1_1.yaml",
    }
    for name, rule_name in (("school", "school_mapping"), ("task", "task_type_rules"), ("channel", "channel_rules"), ("date", "date_rules")):
        assert_snapshot_rule(run_manifest, rule_name, paths[name])
    school = yaml.safe_load(paths["school"].read_text(encoding="utf-8"))
    task = yaml.safe_load(paths["task"].read_text(encoding="utf-8"))
    channel = yaml.safe_load(paths["channel"].read_text(encoding="utf-8"))
    if school.get("status") != "ACTIVE" or school.get("dictionary_version") != "1.1":
        raise RuntimeError("Frozen school dictionary v1.1 is required")
    if task.get("status") != "FROZEN" or task.get("task_type_rules_version") != "17.0":
        raise RuntimeError("Frozen task type rules v17.0 are required")
    if channel.get("status") != "FROZEN" or channel.get("channel_rules_version") != "1.0":
        raise RuntimeError("Frozen channel rules v1.0 are required")
    review_1 = load_json(ROOT / "config/data/school_human_review_round1.json")
    review_2 = load_json(ROOT / "config/data/school_human_review_round2.json")
    if review_1.get("status") != "FROZEN" or review_2.get("status") != "FROZEN":
        raise RuntimeError("Frozen school review mappings are required")
    aliases = {alias: entity for entity in school["canonical_entities"] if entity.get("status") == "ACTIVE" for alias in entity["aliases"]}
    canonical = {entity["canonical_name"]: entity for entity in school["canonical_entities"] if entity.get("status") == "ACTIVE"}
    return {
        "task": task,
        "task_official": set(task["official_task_types"]),
        "task_aliases": {entry["raw_value"]: entry for entry in task["aliases"]},
        "task_multi": {entry["raw_value"]: entry for entry in task["multi_task_rules"]},
        "task_excluded": {entry["raw_value"]: entry for entry in task["excluded_values"]},
        "task_unknown": {entry["raw_value"]: entry for entry in task["unknown_values"]},
        "channel": channel,
        "channel_aliases": {entry["raw_channel"]: entry for entry in channel["aliases"]},
        "school_aliases": aliases,
        "school_canonical": canonical,
        "school_non_school": {entry["original_value"]: entry["classification"] for entry in school.get("non_school_values", []) if entry.get("status") == "ACTIVE"},
        "school_non_university": {entry["original_value"] for entry in school.get("non_university_entities", []) if entry.get("status") == "ACTIVE"},
        "school_unresolved": {entry["original_value"] for entry in school.get("unresolved_values", []) if entry.get("status") == "ACTIVE"} | set(review_1.get("unresolved_without_human_decision", [])),
        "school_review_1": review_1,
        "school_review_2": review_2,
    }


def apply_task_type(record: dict, rules: dict) -> None:
    raw = "" if record.get("task_type") is None else str(record["task_type"]).strip()
    if raw in rules["task_multi"]:
        item = rules["task_multi"][raw]
        status, value, mode, components, rule_id = "MULTI_TASK", "MULTI_TASK", "MULTI_TASK", item["task_type_components"], item["rule_id"]
    elif raw in rules["task_aliases"]:
        item = rules["task_aliases"][raw]
        status, value, mode, components, rule_id = "STANDARDIZED", item["official_task_type"], "SINGLE_TASK", [], item["rule_id"]
    elif raw in rules["task_excluded"]:
        status, value, mode, components, rule_id = "EXCLUDED_BY_BUSINESS_RULE", "", "EXCLUDED", [], "RULE-027"
    elif raw in rules["task_unknown"]:
        status, value, mode, components, rule_id = "UNKNOWN", "UNKNOWN", "UNKNOWN", [], "NO_INFERENCE"
    elif raw in rules["task_official"]:
        status, value, mode, components, rule_id = "STANDARDIZED", raw, "SINGLE_TASK", [], "CANONICAL_IDENTITY"
    else:
        status, value, mode, components, rule_id = "UNMATCHED", "UNKNOWN", "UNKNOWN", [], "NO_FROZEN_RULE_MATCH"
    record.update({"task_type_original": raw, "task_type": value, "task_type_mode": mode, "task_type_components": json.dumps(components, ensure_ascii=False), "task_type_rule_id": rule_id, "task_type_rule_version": "17.0", "task_type_standardization_status": status})


def apply_channel(record: dict, rules: dict) -> None:
    raw = "" if record.get("channel") is None else str(record["channel"]).strip()
    item = rules["channel_aliases"].get(raw)
    if item is None:
        raise RuntimeError(f"No frozen channel v1 rule matches {raw!r}; no inference is permitted")
    record.update({"channel_original": raw, "channel_group": item["channel_group"], "channel": item["channel"] or "", "channel_rule_id": item["rule_id"], "channel_rule_version": "1.0", "channel_standardization_status": "STANDARDIZED"})


def school_id(name: str) -> str:
    return "SCH-" + hashlib.sha1(name.encode("utf-8")).hexdigest()[:10].upper()


def apply_school(record: dict, rules: dict) -> None:
    raw = "" if record.get("school") is None else str(record["school"]).strip()
    country = "" if record.get("country") is None else str(record["country"]).strip()
    final, status, rule_id, version, entity = "", "UNKNOWN", "RULE-013", "1.1", None
    review_2 = rules["school_review_2"].get("approved_mappings", {}).get(raw)
    if review_2:
        final, status, rule_id, version = review_2["canonical_school"], "STANDARDIZED", rules["school_review_2"]["rule_id"], rules["school_review_2"]["version"]
    elif raw in rules["school_review_1"].get("approved_mappings", {}):
        final, status, rule_id, version = rules["school_review_1"]["approved_mappings"][raw], "STANDARDIZED", rules["school_review_1"]["rule_id"], rules["school_review_1"]["version"]
    elif raw in rules["school_review_1"].get("approved_non_school", {}):
        final, status, rule_id, version = "NON_SCHOOL", rules["school_review_1"]["approved_non_school"][raw], rules["school_review_1"]["rule_id"], rules["school_review_1"]["version"]
    elif raw in rules["school_aliases"]:
        entity = rules["school_aliases"][raw]; final, status = entity["canonical_name"], "STANDARDIZED"
    elif raw in rules["school_canonical"]:
        entity = rules["school_canonical"][raw]; final, status = raw, "STANDARDIZED"
    elif raw in rules["school_non_school"]:
        final, status = rules["school_non_school"][raw], rules["school_non_school"][raw]
    elif raw in rules["school_non_university"]:
        final, status = "NON_UNIVERSITY_ENTITY", "NON_UNIVERSITY_ENTITY"
    elif raw in rules["school_unresolved"]:
        final, status = "UNRESOLVED", "UNRESOLVED"
    elif raw:
        final, status = "UNSTANDARDIZED", "UNSTANDARDIZED"
    conflict = ""
    canonical_country = entity.get("canonical_country") if entity else ""
    if canonical_country and country and country != canonical_country:
        conflict = "COUNTRY_SCHOOL_CONFLICT"
    record.update({"school_original": raw, "school": final, "school_id": review_2.get("school_id", school_id(final)) if review_2 else (school_id(final) if status == "STANDARDIZED" else ""), "school_rule_id": rule_id, "school_rule_version": str(version), "school_standardization_status": status, "country_school_conflict": conflict})


def main(run_id: str):
    run_dir = ROOT / "runs" / run_id
    artifacts = run_dir / "artifacts"
    manifest = load_json(artifacts / "source_manifest.json")
    mapping = load_json(artifacts / "schema_mapping.json")
    gate = load_json(run_dir / "gates" / "SCHEMA_GATE.json")
    schema = load_json(run_dir / "snapshots" / "canonical_schema.json")
    run_manifest = load_json(run_dir / "run_manifest.json")
    rules_snapshot = run_dir / "snapshots" / "standardization_rules.yaml"
    business_rules_snapshot = run_dir / "snapshots" / "business_rules.md"
    rules = yaml.safe_load(rules_snapshot.read_text(encoding="utf-8"))
    business_rules = business_rules_snapshot.read_text(encoding="utf-8")
    expected_standardization_rules_version = rules.get("rules_version")
    expected_business_rules_identity, business_rules_identity_type = business_rules_identity(business_rules)

    if gate["decision"] != "PASS":
        raise RuntimeError("Data Standardization requires Schema Gate PASS")
    if schema["status"] != "FROZEN" or schema["schema_version"] != run_manifest["schema_version"]:
        raise RuntimeError("Run snapshot canonical schema must match the frozen manifest version")
    if not expected_standardization_rules_version:
        raise RuntimeError("Run snapshot lacks an authoritative standardization rules version")
    if business_rules_identity_type == 'DECLARED_VERSION' and expected_business_rules_identity != str(expected_standardization_rules_version):
        raise RuntimeError("Run snapshot business rules version does not match standardization rules version")
    if mapping["status"] != "COMPLETED" or manifest["status"] != "COMPLETED":
        raise RuntimeError("Intake and mapping artifacts must be COMPLETED")
    if manifest["received_source_count"] != 6 or len(manifest["sources"]) != 6:
        raise RuntimeError("Expected six received sources")
    dimension_rules = load_dimension_rules(run_manifest)

    canonical_fields = [field["name"] for field in schema["fields"]]
    canonical_set = set(canonical_fields)
    if {"amount", "deadline", "date_original", "date_standardized"} & canonical_set:
        raise RuntimeError("Canonical schema contains forbidden legacy fields")

    mapping_by_source = {}
    for item in mapping["items"]:
        if item["status"] not in ACTIVE_STATUSES:
            raise RuntimeError(f"Mapping item is not eligible for standardization: {item['id']}")
        mapping_by_source[(item["source_id"], item["raw_field_name"])] = item

    standardized_dir = artifacts / "standardized"
    standardized_dir.mkdir(parents=True, exist_ok=True)
    all_rows, source_logs, issues = [], [], []
    global_date_counts, global_degree_counts, global_currency_counts = Counter(), Counter(), Counter()
    expected_excluded = {"跟进反馈", "客户备注", "未成交原因"}

    for source in manifest["sources"]:
        if source["status"] != "RECEIVED":
            raise RuntimeError(f"Unexpected non-received source: {source['source_id']}")
        source_id, year = source["source_id"], source["year"]
        source_path = ROOT / source["file_path"]
        wb = load_workbook(source_path, read_only=True, data_only=True)
        sheet_name = wb.sheetnames[0]
        ws = wb[sheet_name]
        headers = [cell.value.strip() if isinstance(cell.value, str) else cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        profile = load_json(artifacts / "source_profiles" / f"{source_id}.json")
        profile_fields = set(profile["sheets"][sheet_name]["fields"][idx]["field_name"] for idx in range(len(headers)))
        if set(headers) != profile_fields:
            raise RuntimeError(f"Raw headers diverge from source profile for {source_id}")

        output_rows, blank_rows, date_counts, degree_counts, currency_counts = [], 0, Counter(), Counter(), Counter()
        excluded_fields = [header for header in headers if mapping_by_source[(source_id, header)]["status"] == "EXCLUDED_BY_BUSINESS_RULE"]
        if not set(excluded_fields).issubset(expected_excluded):
            raise RuntimeError(f"Unexpected excluded field in {source_id}")

        for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            raw = dict(zip(headers, values))
            if all(empty(value) for value in raw.values()):
                blank_rows += 1
                continue
            record = {field: None for field in canonical_fields}
            record.update({
                "source_id": source_id,
                "source_file": source_path.name,
                "source_sheet": sheet_name,
                "source_row_id": excel_row,
                "year": year,
                "department": source["department"],
            })
            for header, value in raw.items():
                item = mapping_by_source[(source_id, header)]
                if item["status"] == "EXCLUDED_BY_BUSINESS_RULE":
                    continue
                target = item["canonical_field"]
                if target == "consultation_date":
                    record["consultation_date_original"] = value.isoformat() if isinstance(value, (dt.date, dt.datetime)) else value
                    standardized, kind, review = iso_date(value, year)
                    record["consultation_date"] = standardized
                    date_counts[f"consultation_date:{kind or 'review_required'}"] += 1
                    if review:
                        row_issue(issues, source_id, excel_row, "consultation_date", value, review, "RULE-002~RULE-006")
                elif target == "deadline":
                    record["ddl_original"] = value.isoformat() if isinstance(value, (dt.date, dt.datetime)) else value
                    standardized, kind, review = iso_date(value, year)
                    record["ddl"] = standardized
                    date_counts[f"ddl:{kind or 'review_required'}"] += 1
                    if review:
                        row_issue(issues, source_id, excel_row, "ddl", value, review, "RULE-002~RULE-006")
                elif target == "degree_level":
                    record["degree_level_original"] = value
                    if empty(value):
                        record["degree_level"] = None
                        degree_counts["blank"] += 1
                    else:
                        normalized = DEGREE_MAP.get(str(value).strip(), "UNKNOWN")
                        record["degree_level"] = normalized
                        degree_counts[f"{str(value).strip()}->{normalized}"] += 1
                elif target == "amount":
                    amount, review = decimal_value(value)
                    if review:
                        row_issue(issues, source_id, excel_row, "amount_original", value, review, item.get("value_rule_ref"))
                        currency_counts["review_required"] += 1
                    elif amount is not None:
                        rule = item.get("value_rule_ref")
                        currency = "AUD" if rule == "RULE-009" else "CNY" if rule == "RULE-010" else None
                        if currency is None:
                            raise RuntimeError(f"Amount item lacks a currency rule: {item['id']}")
                        record["amount_original"] = float(amount)
                        record["currency_original"] = currency
                        converted = (amount * Decimal("4.5")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP) if currency == "AUD" else amount
                        record["amount_cny"] = float(converted)
                        currency_counts[f"{currency}_records"] += 1
                else:
                    if target not in canonical_set:
                        raise RuntimeError(f"Mapping target is not canonical: {target}")
                    record[target] = value
            apply_school(record, dimension_rules)
            apply_task_type(record, dimension_rules)
            apply_channel(record, dimension_rules)
            output_rows.append(record)

        wb.close()
        source_csv = standardized_dir / f"{source_id}_standardized.csv"
        write_csv(source_csv, canonical_fields, output_rows)
        all_rows.extend(output_rows)
        source_logs.append({
            "source_id": source_id, "input_rows": len(list(load_workbook(source_path, read_only=True, data_only=True)[sheet_name].iter_rows(min_row=2, values_only=True))),
            "blank_rows_excluded": blank_rows, "output_rows": len(output_rows),
            "field_mappings": {"mapped_fields": len(headers) - len(excluded_fields), "excluded_fields": excluded_fields},
            "date_transformations": dict(date_counts), "degree_transformations": dict(degree_counts),
            "currency_transformations": dict(currency_counts), "excluded_fields": excluded_fields,
            "unknown_values": sum(count for key, count in degree_counts.items() if key.endswith("->UNKNOWN")),
            "review_required_values": sum(1 for issue in issues if issue["source_id"] == source_id),
            "rule_ids": ["RULE-001", "RULE-002", "RULE-003", "RULE-004", "RULE-005", "RULE-006", "RULE-007", "RULE-008", "RULE-009", "RULE-010", "RULE-011", "RULE-012"]
        })
        global_date_counts.update(date_counts); global_degree_counts.update(degree_counts); global_currency_counts.update(currency_counts)

    write_csv(artifacts / "unified_dataset.csv", canonical_fields, all_rows)
    write_xlsx(artifacts / "unified_dataset.xlsx", canonical_fields, all_rows)
    review = {"run_id": run_id, "agent": "Data Standardization Agent", "status": "HUMAN_REVIEW_REQUIRED" if issues else "COMPLETED", "review_required_count": len(issues), "issues": issues}
    write_json(artifacts / "standardization_review.json", review)
    cleaning = {
        "run_id": run_id, "agent": "Data Standardization Agent", "wave": "W3", "version": 1,
        "status": review["status"], "business_rules_version": expected_standardization_rules_version, "business_rules_identity": expected_business_rules_identity, "business_rules_identity_type": business_rules_identity_type, "business_rules_checksum": checksum(business_rules_snapshot), "standardization_rules_checksum": checksum(rules_snapshot), "canonical_schema_version": schema["schema_version"],
        "inputs": ["source_manifest.json", "source_profiles/*.json", "schema_mapping.json", "gates/SCHEMA_GATE.json", "snapshots/canonical_schema.json", "snapshots/business_rules.md", "snapshots/standardization_rules.yaml"],
        "sources": source_logs,
        "totals": {"input_rows": sum(item["input_rows"] for item in source_logs), "blank_rows_excluded": sum(item["blank_rows_excluded"] for item in source_logs), "output_rows": len(all_rows), "date_transformations": dict(global_date_counts), "degree_transformations": dict(global_degree_counts), "currency_transformations": dict(global_currency_counts), "review_required_count": len(issues)},
        "excluded_fields": sorted(expected_excluded), "unified_dataset_fields": canonical_fields,
        "evidence_refs": [{"type": "manual_business_confirmation", "rule_id": f"RULE-{n:03d}"} for n in range(1, 13)]
    }
    write_json(artifacts / "cleaning_log.json", cleaning)
    standardization_rules = {"run_id": run_id, "agent": "Data Standardization Agent", "wave": "W3", "version": 1, "status": review["status"], "generated_at": dt.datetime.now(dt.timezone.utc).isoformat(), "items": source_logs, "unresolved": issues, "evidence_refs": cleaning["evidence_refs"]}
    write_json(artifacts / "standardization_rules.json", standardization_rules)
    # A05 owns the run-bound task/channel standardization evidence consumed by A06.
    # This helper is read-only against the newly written unified dataset.
    from complete_a05_standardization_audits_v1 import complete_a05_audits
    complete_a05_audits(run_id)
    print(json.dumps({"status": review["status"], "sources": source_logs, "output_rows": len(all_rows), "review_required_count": len(issues)}, ensure_ascii=False, indent=2))


def write_csv(path: Path, fields, rows):
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields, extrasaction="raise")
        writer.writeheader(); writer.writerows(rows)


def write_xlsx(path: Path, fields, rows):
    wb = Workbook(write_only=True); ws = wb.create_sheet("unified_dataset"); ws.append(fields)
    for row in rows: ws.append([row[field] for field in fields])
    wb.save(path)


def write_json(path: Path, value):
    path.write_text(json.dumps(value, ensure_ascii=False, indent=2), encoding="utf-8")


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else RUN_ID)
