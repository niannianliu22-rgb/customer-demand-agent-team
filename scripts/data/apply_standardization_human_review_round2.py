#!/usr/bin/env python3
"""Apply the three explicit School Human Review Round 2 approvals only."""
from __future__ import annotations
import csv, hashlib, json
from datetime import datetime, timezone
from pathlib import Path
from openpyxl import load_workbook

ROOT=Path(__file__).resolve().parents[2]; RUN_ID='RUN-202608-DEMAND-001'
ART=ROOT/f'runs/{RUN_ID}/artifacts'; DATA=ART/'unified_dataset.csv'; XLSX=ART/'unified_dataset.xlsx'
RULE=ROOT/'config/data/school_human_review_round2.json'; SCHEMA=ROOT/'schemas/canonical_schema.json'; OUT=ROOT/'quality'
def sha(p): return hashlib.sha256(p.read_bytes()).hexdigest()
def read(p):
    with p.open(encoding='utf-8-sig',newline='') as h:return list(csv.DictReader(h))
def write(p,fields,rows):
    with p.open('w',encoding='utf-8-sig',newline='') as h:
        w=csv.DictWriter(h,fieldnames=fields);w.writeheader();w.writerows(rows)
def main():
    before={str(p.relative_to(ROOT)):sha(p) for p in [DATA,XLSX,RULE,SCHEMA]}
    rule=json.loads(RULE.read_text(encoding='utf-8'))
    rows=read(DATA); fields=list(rows[0])
    if 'school_id' not in fields: fields.append('school_id')
    applied=[]
    for r in rows:
        mapping=rule['approved_mappings'].get(r['school_original'])
        if not mapping: continue
        prior=r['school']; r['school']=mapping['canonical_school']; r['school_id']=mapping['school_id']; r['school_rule_id']=rule['rule_id']; r['school_rule_version']=rule['version']; r['school_standardization_status']='STANDARDIZED'
        applied.append({'record_id':f"{r['source_id']}:{r['source_row_id']}",'raw_school':r['school_original'],'previous_school':prior,'canonical_school':r['school'],'school_id':r['school_id'],'manual_decision':'MAP_TO_CANONICAL','review_status':'APPROVED','rule_id':rule['rule_id'],'rule_version':rule['version']})
    write(DATA,fields,rows)
    wb=load_workbook(XLSX); ws=wb.active; old=[c.value for c in ws[1]]
    for col,name in enumerate(fields,1):ws.cell(1,col).value=name
    for rn,r in enumerate(rows,2):
        for col,name in enumerate(fields,1):ws.cell(rn,col).value=r.get(name,'')
    for col in range(len(fields)+1,len(old)+1):
        for rn in range(1,ws.max_row+1):ws.cell(rn,col).value=None
    wb.save(XLSX)
    OUT.mkdir(exist_ok=True)
    out_fields=['record_id','raw_school','previous_school','canonical_school','school_id','manual_decision','review_status','rule_id','rule_version']
    write(OUT/'school_human_review_round2_applied.csv',out_fields,applied)
    audit={'run_id':RUN_ID,'artifact':'standardization_human_review_round2_application','applied_at':datetime.now(timezone.utc).isoformat(),'applied_rows':len(applied),'rule':{'path':str(RULE.relative_to(ROOT)),'rule_id':rule['rule_id'],'version':rule['version'],'status':rule['status']},'schema_version':'1.5.0','before_checksums':before,'after_checksums':{str(p.relative_to(ROOT)):sha(p) for p in [DATA,XLSX,SCHEMA]},'not_modified':['Date','Task Type','Channel','Historical Demand','Academic Calendar','Business Rules']}
    (OUT/'standardization_human_review_round2_application_audit.json').write_text(json.dumps(audit,ensure_ascii=False,indent=2),encoding='utf-8')
    print(json.dumps({'applied_rows':len(applied),'school_blockers_expected':sum(r['school']=='UNSTANDARDIZED' for r in rows)},ensure_ascii=False))
if __name__=='__main__':main()
