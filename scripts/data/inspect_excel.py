"""
inspect_excel.py — deterministic Excel structure scanner.

This is the ONLY tool Data Intake Agent is allowed to use to learn the real
structure of an input file. It performs no business-semantic interpretation:
it reports what is literally in the file (sheet names, row/column counts,
field names, dtypes, missing rates, sample values, and structural date/amount
*candidate* flags based on naming/type heuristics only — never a mapping to
a canonical business field). Semantic mapping is Schema Mapping Agent's job.

Read-only: never writes to, or modifies, the inspected file.
"""

import datetime
import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:  # pragma: no cover
    print(json.dumps({"error": f"openpyxl not available: {exc}"}))
    sys.exit(1)

DATE_NAME_HINTS = ["日期", "时间", "date", "time"]
AMOUNT_NAME_HINTS = ["金额", "费用", "价格", "单价", "学费", "amount", "price", "fee"]


def _classify_value(value):
    if value is None or value == "":
        return "empty"
    if isinstance(value, (datetime.datetime, datetime.date)):
        return "datetime"
    if isinstance(value, bool):
        return "bool"
    if isinstance(value, int):
        return "int"
    if isinstance(value, float):
        return "float"
    return "string"


def _sample(value):
    if isinstance(value, (datetime.datetime, datetime.date)):
        return value.isoformat()
    return value


def _inspect_sheet(ws):
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0

    # Empty sheet: no rows, or only a header row with zero data rows.
    if max_row == 0 or max_col == 0:
        return {
            "row_count": 0,
            "column_count": 0,
            "is_empty_sheet": True,
            "fields": [],
            "date_candidate_fields": [],
            "amount_candidate_fields": [],
        }

    rows = list(ws.iter_rows(min_row=1, max_row=max_row, max_col=max_col, values_only=True))
    header = [str(c).strip() if c is not None else f"__unnamed_col_{i+1}__" for i, c in enumerate(rows[0])]
    data_rows = rows[1:]
    row_count = len(data_rows)
    blank_row_count = sum(1 for r in data_rows if all(c is None for c in r))

    is_empty_sheet = row_count == 0 or blank_row_count == row_count

    fields = []
    date_candidates = []
    amount_candidates = []

    for col_idx, field_name in enumerate(header):
        col_values = [r[col_idx] if col_idx < len(r) else None for r in data_rows]
        non_null = [v for v in col_values if v is not None and v != ""]
        missing_count = row_count - len(non_null)
        missing_rate = round(missing_count / row_count, 4) if row_count > 0 else None

        type_counts = {}
        for v in non_null:
            t = _classify_value(v)
            type_counts[t] = type_counts.get(t, 0) + 1

        if not non_null:
            dtype = "empty"
        elif len(type_counts) == 1:
            dtype = next(iter(type_counts))
        else:
            dtype = "mixed(" + ",".join(sorted(type_counts)) + ")"

        sample_values = [_sample(v) for v in non_null[:3]]

        name_lower = field_name.lower()
        is_date_by_name = any(h in field_name or h in name_lower for h in DATE_NAME_HINTS)
        is_date_by_type = dtype == "datetime"
        is_amount_by_name = any(h in field_name or h in name_lower for h in AMOUNT_NAME_HINTS)
        is_amount_by_type = dtype in ("int", "float") and not is_date_by_type

        field_profile = {
            "field_name": field_name,
            "dtype": dtype,
            "missing_rate": missing_rate,
            "sample_values": sample_values,
        }
        fields.append(field_profile)

        if is_date_by_name or is_date_by_type:
            date_candidates.append(field_name)
        if is_amount_by_name or (is_amount_by_type and is_amount_by_name is False and False):
            # amount-by-type alone is too weak a signal (e.g. counts, IDs);
            # only flag when the column name also hints at an amount.
            pass
        if is_amount_by_name:
            amount_candidates.append(field_name)

    return {
        "row_count": row_count,
        "blank_row_count": blank_row_count,
        "column_count": len(header),
        "is_empty_sheet": is_empty_sheet,
        "fields": fields,
        "date_candidate_fields": date_candidates,
        "amount_candidate_fields": amount_candidates,
    }


def inspect_excel(file_path: str) -> dict:
    """Deterministically scan one Excel file. Never raises: read errors are
    captured in the returned dict's `errors` list."""
    path = Path(file_path)
    result = {
        "file_name": path.name,
        "file_path": str(file_path),
        "file_format": path.suffix.lstrip(".").lower() or "unknown",
        "readable": False,
        "sheet_names": [],
        "sheets": {},
        "warnings": [],
        "errors": [],
    }

    if not path.exists():
        result["errors"].append(f"file does not exist: {file_path}")
        return result

    try:
        wb = openpyxl.load_workbook(filename=str(path), read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — must capture, never crash the scan
        result["errors"].append(f"{type(exc).__name__}: {exc}")
        return result

    try:
        result["sheet_names"] = list(wb.sheetnames)
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            try:
                profile = _inspect_sheet(ws)
            except Exception as exc:  # noqa: BLE001
                profile = {"error": f"{type(exc).__name__}: {exc}"}
                result["warnings"].append(f"sheet '{sheet_name}' could not be fully profiled: {exc}")
            result["sheets"][sheet_name] = profile
            if profile.get("is_empty_sheet"):
                result["warnings"].append(f"sheet '{sheet_name}' is empty")
            elif profile.get("blank_row_count", 0) > 0:
                result["warnings"].append(
                    f"sheet '{sheet_name}' has {profile['blank_row_count']} fully-blank row(s) "
                    f"within its used range (row_count={profile['row_count']}); "
                    "likely trailing blank rows inflating the sheet's used range"
                )
        result["readable"] = True
    finally:
        wb.close()

    return result


if __name__ == "__main__":
    if len(sys.argv) != 2:
        print(json.dumps({"error": "usage: inspect_excel.py <file_path>"}))
        sys.exit(1)
    print(json.dumps(inspect_excel(sys.argv[1]), ensure_ascii=False, indent=2))
