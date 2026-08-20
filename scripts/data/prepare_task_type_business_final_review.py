#!/usr/bin/env python3
"""Create a concise, evidence-first business final-review queue for task types."""
from __future__ import annotations

import csv
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
RUN = ROOT / "runs/RUN-202608-DEMAND-001"
DIM = RUN / "artifacts/dimension_review"
ROUND3 = DIM / "task_type_official_mapping_review_round3.csv"
CLUSTERS = DIM / "task_type_cluster_review_round4.md"
MULTI = DIM / "task_type_multi_task_review_round4.csv"
EXCEPTIONS = DIM / "task_type_exception_review_round4.csv"
DATASET = RUN / "artifacts/unified_dataset.csv"


def string(value: object) -> str:
    return value.isoformat() if isinstance(value, (datetime, date)) else str(value or "")


def parse_cluster_flags() -> dict[str, str]:
    flags = {}; current = None
    for line in CLUSTERS.read_text(encoding="utf-8").splitlines():
        if line.startswith("## "): current = line[3:]
        elif current and line.startswith("Cluster risk: `"):
            flags[current] = line.removeprefix("Cluster risk: `").removesuffix("`")
    return flags


def raw_context(scope: set[str]) -> dict[str, dict[str, list[str]]]:
    with DATASET.open(encoding="utf-8-sig", newline="") as handle: rows=list(csv.DictReader(handle))
    books = {}
    for record in rows:
        raw=(record.get("task_type") or "").strip()
        if raw not in scope or record["source_file"] in books: continue
        wb=load_workbook(RUN/"input"/record["source_file"],read_only=True,data_only=False)
        ws=wb[record["source_sheet"]]; values=list(ws.iter_rows(values_only=True)); wb.close()
        books[record["source_file"]]=(values[0],values)
    grouped=defaultdict(lambda: defaultdict(list))
    for record in rows:
        raw=(record.get("task_type") or "").strip()
        if raw not in scope or len(grouped[raw]["sample_context"])>=3: continue
        headers, values=books[record["source_file"]]
        original={str(key):value for key,value in zip(headers,values[int(record["source_row_id"])-1])}
        task_field="作业形式" if "作业形式" in original else ("作业类型" if "作业类型" in original else "咨询内容")
        course=original.get("专业/课程",original.get("专业",""))
        notes=" | ".join(string(original[key]) for key in ("跟进反馈","客户备注","未成交原因") if original.get(key) not in (None,"","/"))
        order="; ".join(f"{key}={string(original[key])}" for key in ("订单编号","进度") if key in original and original[key] not in (None,"","/"))
        product="NOT_AVAILABLE_IN_SOURCE"
        grouped[raw]["original_order_type"].append(f"{task_field}={string(original.get(task_field))}")
        grouped[raw]["product_name"].append(product)
        grouped[raw]["service_description"].append(string(original.get(task_field)))
        grouped[raw]["customer_requirement"].append(string(course))
        grouped[raw]["order_note"].append((notes + (" | " if notes and order else "") + order) or "NOT_AVAILABLE")
        grouped[raw]["sample_context"].append(f"{record['source_id']}:{record['source_row_id']} [原始{task_field}={string(original.get(task_field))}; 课程/专业={string(course) or '—'}; 备注/订单={notes or '—'} {order}]")
    out={}
    for raw,parts in grouped.items():
        out[raw]={key:" || ".join(dict.fromkeys(values)) for key,values in parts.items()}
    return out


def main() -> None:
    with ROUND3.open(encoding="utf-8-sig",newline="") as f: rows=list(csv.DictReader(f))
    with MULTI.open(encoding="utf-8-sig",newline="") as f: multi={r["raw_value"]:r for r in csv.DictReader(f)}
    with EXCEPTIONS.open(encoding="utf-8-sig",newline="") as f: exceptions={r["raw_value"]:r for r in csv.DictReader(f)}
    flags=parse_cluster_flags()
    semantic_clusters={name for name,flag in flags.items() if "SEMANTIC_CONFLICT" in flag}
    selected=[]
    for row in rows:
        canonical=row["proposed_official_task_type"]
        risk=flags.get(canonical,"NONE") if canonical else "NONE"
        if canonical in semantic_clusters or row["classification"] in {"PROPOSED_MEDIUM","REVIEW_REQUIRED","MULTI_TASK","NON_TASK","UNKNOWN"}:
            selected.append({**row,"risk_flag":risk})
    scope={row["raw_value"] for row in selected}
    contexts=raw_context(scope)
    priorities={"RISK_CLUSTER":1,"REVIEW_REQUIRED":2,"PROPOSED_MEDIUM":3,"MULTI_TASK":4,"EXCEPTION":5}
    final=[]
    for row in selected:
        raw=row["raw_value"]; classification=row["classification"]; canonical=row["proposed_official_task_type"]
        if canonical in semantic_clusters: bucket="RISK_CLUSTER"
        elif classification=="REVIEW_REQUIRED": bucket="REVIEW_REQUIRED"
        elif classification=="PROPOSED_MEDIUM": bucket="PROPOSED_MEDIUM"
        elif classification=="MULTI_TASK": bucket="MULTI_TASK"
        else: bucket="EXCEPTION"
        ctx=contexts.get(raw,{})
        multi_row=multi.get(raw,{})
        exception=exceptions.get(raw,{})
        if classification=="MULTI_TASK":
            recommendation="MULTI_TASK_REQUIRES_PRIMARY_AND_LIST"
            reason="保留多个任务语义，不强制映射为单一官方订单类型。"
            insufficient="INSUFFICIENT_EVIDENCE_FOR_PRIMARY" if not multi_row.get("recommended_primary_task_type") else "EVIDENCE_AVAILABLE"
        elif classification in {"REVIEW_REQUIRED","UNKNOWN"}:
            recommendation="KEEP_REVIEW_REQUIRED"
            reason="INSUFFICIENT_EVIDENCE：原始订单上下文不足以唯一落入公司官方订单类型。"
            insufficient="INSUFFICIENT_EVIDENCE"
        elif classification=="NON_TASK":
            recommendation="KEEP_NON_TASK"
            reason=f"疑似非任务元数据，应优先核实 {exception.get('possible_other_field','other_field')}。"
            insufficient="NOT_APPLICABLE"
        else:
            recommendation=canonical
            reason="存在官方范围内的候选，但需人工确认以消除聚类风险。"
            insufficient="EVIDENCE_AVAILABLE" if ctx.get("customer_requirement") else "INSUFFICIENT_EVIDENCE"
        final.append({
            "_priority":priorities[bucket], "_bucket":bucket,
            "raw_value":raw,"record_count":row["record_count"],"current_classification":classification,
            "proposed_official_task_type":canonical,"risk_flag":row["risk_flag"],
            "original_order_type":ctx.get("original_order_type","NOT_AVAILABLE_IN_SOURCE"),"product_name":ctx.get("product_name","NOT_AVAILABLE_IN_SOURCE"),
            "service_description":ctx.get("service_description","NOT_AVAILABLE_IN_SOURCE"),"customer_requirement":ctx.get("customer_requirement","NOT_AVAILABLE_IN_SOURCE"),
            "order_note":ctx.get("order_note","NOT_AVAILABLE_IN_SOURCE"),"sample_context":ctx.get("sample_context",row["sample_context"]),
            "model_recommendation":recommendation,"model_reason":f"{reason} {row['evidence']}","evidence_status":insufficient,
            "recommended_primary_task_type":multi_row.get("recommended_primary_task_type",""),"recommended_task_type_list":multi_row.get("recommended_task_type_list",""),
            "exception_suspected_type":exception.get("possible_other_field",""),"business_decision":"","final_official_task_type":"","review_note":"",
        })
    final.sort(key=lambda row:(row["_priority"],-int(row["record_count"]),row["raw_value"]))
    fields=["raw_value","record_count","current_classification","proposed_official_task_type","risk_flag","original_order_type","product_name","service_description","customer_requirement","order_note","sample_context","model_recommendation","model_reason","evidence_status","recommended_primary_task_type","recommended_task_type_list","exception_suspected_type","business_decision","final_official_task_type","review_note"]
    with (DIM/"task_type_business_final_review.csv").open("w",encoding="utf-8",newline="") as f:
        writer=csv.DictWriter(f,fieldnames=fields);writer.writeheader();writer.writerows([{key:row[key] for key in fields} for row in final])
    direct=sum(
        row["classification"] in {"EXACT_MATCH", "PROPOSED_HIGH"}
        and flags.get(row["proposed_official_task_type"], "NONE") == "NONE"
        for row in rows
    )
    human=len(final)
    insufficient=sum(row["evidence_status"].startswith("INSUFFICIENT_EVIDENCE") for row in final)
    summary=f"""# Task Type Business Final Review Summary

This is a human-decision queue only. No unified dataset, official canonical list, or aliases file was changed.

| Metric | Unique raw values |
|---|---:|
| Directly acceptable candidates outside all cluster-risk flags | {direct} |
| Requires business confirmation in this final-review queue | {human} |
| Insufficient evidence | {insufficient} |
| MULTI_TASK | {sum(r['current_classification']=='MULTI_TASK' for r in final)} |
| NON_TASK / UNKNOWN | {sum(r['current_classification'] in {'NON_TASK','UNKNOWN'} for r in final)} |
| Highest-priority semantic-conflict clusters | {len(semantic_clusters)} |

Priority ordering in the CSV: semantic-risk clusters → REVIEW_REQUIRED → PROPOSED_MEDIUM → MULTI_TASK → NON_TASK / UNKNOWN. `final_official_task_type` is intentionally blank and must later be selected only from the 66 company official order types.
"""
    (DIM/"task_type_business_final_review_summary.md").write_text(summary,encoding="utf-8")
    if len(final) != len({r['raw_value'] for r in final}): raise ValueError("business final-review queue contains duplicate raw values")
    print({"directly_acceptable":direct,"needs_human":human,"insufficient_evidence":insufficient,"multi_task":sum(r['current_classification']=='MULTI_TASK' for r in final),"non_task_unknown":sum(r['current_classification'] in {'NON_TASK','UNKNOWN'} for r in final)})


if __name__=="__main__": main()
