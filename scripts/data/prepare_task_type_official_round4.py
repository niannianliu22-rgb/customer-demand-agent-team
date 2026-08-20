#!/usr/bin/env python3
"""Create Round-4 human review artifacts for official task-type candidates."""
from __future__ import annotations

import csv
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
RUN = ROOT / "runs/RUN-202608-DEMAND-001"
DIMENSION = RUN / "artifacts/dimension_review"
CANONICAL = ROOT / "config/dimensions/task_type/canonical.csv"
ALIASES = ROOT / "config/dimensions/task_type/aliases_candidate.csv"
ROUND3 = DIMENSION / "task_type_official_mapping_review_round3.csv"
DATASET = RUN / "artifacts/unified_dataset.csv"


def value_text(value: object) -> str:
    return value.isoformat() if isinstance(value, (datetime, date)) else str(value or "")


def direct_context(scope: set[str]) -> dict[str, list[str]]:
    with DATASET.open(encoding="utf-8-sig", newline="") as handle:
        data = list(csv.DictReader(handle))
    caches = {}
    for row in data:
        raw = (row.get("task_type") or "").strip()
        if raw not in scope or row["source_file"] in caches:
            continue
        workbook = load_workbook(RUN / "input" / row["source_file"], read_only=True, data_only=False)
        sheet = workbook[row["source_sheet"]]
        workbook_rows = list(sheet.iter_rows(values_only=True))
        workbook.close()
        caches[row["source_file"]] = (workbook_rows[0], workbook_rows)
    results = defaultdict(list)
    for row in data:
        raw = (row.get("task_type") or "").strip()
        if raw not in scope or len(results[raw]) >= 3:
            continue
        headers, rows = caches[row["source_file"]]
        original = {str(key): item for key, item in zip(headers, rows[int(row["source_row_id"]) - 1])}
        task_field = "作业形式" if "作业形式" in original else ("作业类型" if "作业类型" in original else "咨询内容")
        course = original.get("专业/课程", original.get("专业", ""))
        notes = " | ".join(value_text(original[key]) for key in ("跟进反馈", "客户备注", "未成交原因") if original.get(key) not in (None, "", "/"))
        results[raw].append(f"{row['source_id']}:{row['source_row_id']} [原始{task_field}={value_text(original.get(task_field))}; 课程/专业={value_text(course) or '—'}; 服务/备注={notes or '—'}]")
    return results


def multi_officials(raw: str) -> list[str]:
    # Every retained value comes from canonical.csv. Unmatched components are
    # documented in evidence rather than represented as invented types.
    lower = raw.casefold()
    found = []
    def add(name: str) -> None:
        if name not in found: found.append(name)
    if "essay" in lower: add("essay")
    if "润色" in raw: add("润色-proofreading")
    if "report" in lower: add("report")
    if "海报" in raw: add("海报")
    if "ppt" in lower or "pre" in lower or "演讲稿" in raw: add("presentation")
    if "考试" in raw: add("考试")
    if "补考" in raw: pass  # internal/external cannot be chosen without evidence
    if "毕业论文" in raw or "大论文" in raw: add("Dissertation")
    if "选课" in raw: add("选课")
    if "入学测试" in raw: pass
    if "数据" in raw: add("Analysis")
    return found


def cluster_flags(canonical: str, rows: list[dict[str, str]]) -> list[str]:
    flags = []
    raw_values = [row["raw_value"] for row in rows]
    if any(row["classification"] == "PROPOSED_MEDIUM" for row in rows): flags.append("POSSIBLE_OVER_MERGE")
    if canonical == "Dissertation" and any(any(token in raw for token in ("全包", "半包", "数据", "润色", "辅导")) for raw in raw_values): flags.append("POSSIBLE_OVER_MERGE")
    if canonical == "润色-proofreading" and any("论文" in raw for raw in raw_values): flags.append("POSSIBLE_WRONG_MAPPING")
    if canonical in {"presentation", "assignment", "tutoring", "course work", "Paper"} and len(set(raw_values)) > 1: flags.append("SEMANTIC_CONFLICT")
    return flags or ["NONE"]


def main() -> None:
    with CANONICAL.open(encoding="utf-8-sig", newline="") as handle: official = list(csv.DictReader(handle))
    with ALIASES.open(encoding="utf-8-sig", newline="") as handle: aliases = list(csv.DictReader(handle))
    with ROUND3.open(encoding="utf-8-sig", newline="") as handle: round3 = list(csv.DictReader(handle))
    if len(official) != 66 or len(aliases) != 210 or len(round3) != 210: raise ValueError("official Round-3 inputs are incomplete")
    review = {row["raw_value"]: row for row in round3}
    if set(review) != {row["raw_value"] for row in aliases}: raise ValueError("candidate and review raw-value coverage differs")
    manual_rows = [row for row in round3 if row["classification"] in {"PROPOSED_MEDIUM", "REVIEW_REQUIRED", "UNKNOWN"}]
    multi_rows = [row for row in round3 if row["classification"] == "MULTI_TASK"]
    exception_rows = [row for row in round3 if row["classification"] in {"NON_TASK", "UNKNOWN"}]
    context_scope = {row["raw_value"] for row in manual_rows + multi_rows + exception_rows}
    raw_context = direct_context(context_scope)
    def sample(raw: str) -> str: return " || ".join(raw_context.get(raw, [])) or review[raw]["sample_context"]

    # Official cluster report: only proposed/exact mappings with a real official target.
    clusters = defaultdict(list)
    for row in round3:
        if row["classification"] in {"EXACT_MATCH", "PROPOSED_HIGH", "PROPOSED_MEDIUM"} and row["proposed_official_task_type"]:
            clusters[row["proposed_official_task_type"]].append(row)
    accepted = conflict = 0
    sections = ["# Task Type Official Cluster Review — Round 4", "", "All candidates below use the 66 values in `config/dimensions/task_type/canonical.csv`. `NONE` means no cluster-level semantic warning was detected; it is not an automatic writeback approval.", ""]
    for canonical, items in sorted(clusters.items()):
        flags = cluster_flags(canonical, items)
        if flags == ["NONE"]: accepted += 1
        else: conflict += 1
        sections.extend([f"## {canonical}", "", f"Cluster risk: `{', '.join(flags)}`", "", "| raw_value | record_count | classification | confidence | sample_context | evidence |", "|---|---:|---|---|---|---|"])
        for item in items:
            context = sample(item["raw_value"]).replace("|", "\\|")
            evidence = item["evidence"].replace("|", "\\|")
            sections.append(f"| {item['raw_value'] or '<BLANK>'} | {item['record_count']} | {item['classification']} | {item['confidence']} | {context} | {evidence} |")
        sections.append("")
    (DIMENSION / "task_type_cluster_review_round4.md").write_text("\n".join(sections), encoding="utf-8")

    manual_fields = ["raw_value","record_count","classification","proposed_official_task_type","confidence","sample_context","evidence","recommended_official_task_type","decision","final_official_task_type","review_note"]
    with (DIMENSION / "task_type_manual_review_round4.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=manual_fields); writer.writeheader()
        for item in manual_rows:
            # A proposed target is a recommendation only when previous confidence was MEDIUM.
            recommendation = item["proposed_official_task_type"] if item["classification"] == "PROPOSED_MEDIUM" else ""
            writer.writerow({"raw_value":item["raw_value"],"record_count":item["record_count"],"classification":item["classification"],"proposed_official_task_type":item["proposed_official_task_type"],"confidence":item["confidence"],"sample_context":sample(item["raw_value"]),"evidence":item["evidence"],"recommended_official_task_type":recommendation,"decision":"","final_official_task_type":"","review_note":""})
    multi_fields = ["raw_value","record_count","detected_official_task_types","recommended_primary_task_type","recommended_task_type_list","confidence","evidence","decision","review_note"]
    with (DIMENSION / "task_type_multi_task_review_round4.csv").open("w", encoding="utf-8", newline="") as handle:
        writer=csv.DictWriter(handle,fieldnames=multi_fields); writer.writeheader()
        for item in multi_rows:
            tasks=multi_officials(item["raw_value"])
            writer.writerow({"raw_value":item["raw_value"],"record_count":item["record_count"],"detected_official_task_types":"|".join(tasks),"recommended_primary_task_type":"","recommended_task_type_list":"|".join(tasks),"confidence":"MULTI_TASK","evidence":f"{item['evidence']} 原始 Excel：{sample(item['raw_value'])}","decision":"","review_note":""})
    exception_fields = ["raw_value","record_count","classification","sample_context","evidence","is_true_exception","possible_other_field","recommended_official_task_type","decision","review_note"]
    other_fields = {"预存":"payment_transaction","毕业无忧定金":"payment_stage","SVIP预存":"payment_transaction","vip充值":"payment_transaction","包课补款":"payment_transaction","预存升级学年包定金":"payment_stage","/":"missing_value","":"missing_value"}
    with (DIMENSION / "task_type_exception_review_round4.csv").open("w", encoding="utf-8", newline="") as handle:
        writer=csv.DictWriter(handle,fieldnames=exception_fields); writer.writeheader()
        for item in exception_rows:
            writer.writerow({"raw_value":item["raw_value"],"record_count":item["record_count"],"classification":item["classification"],"sample_context":sample(item["raw_value"]),"evidence":item["evidence"],"is_true_exception":"YES","possible_other_field":other_fields.get(item["raw_value"],"missing_or_transaction_metadata"),"recommended_official_task_type":"","decision":"","review_note":""})
    if len(manual_rows) != 131 or len(multi_rows) != 20 or len(exception_rows) != 8: raise ValueError("Round-4 category counts changed")
    print({"directly_acceptable_clusters":accepted,"conflict_clusters":conflict,"proposed_medium":sum(item['classification']=='PROPOSED_MEDIUM' for item in round3),"review_required":sum(item['classification']=='REVIEW_REQUIRED' for item in round3),"multi_task":len(multi_rows),"non_task_unknown":len(exception_rows)})


if __name__ == "__main__": main()
