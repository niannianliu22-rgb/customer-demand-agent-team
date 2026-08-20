#!/usr/bin/env python3
"""Map historical task_type values to the company official order-type range.

Proposal-only: no unified dataset or official dictionary is changed.
"""
from __future__ import annotations

import csv
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
RUN = ROOT / "runs/RUN-202608-DEMAND-001"
OFFICIAL_FILE = ROOT / "config/dimensions/task_type/official_order_types.xlsx"
CANONICAL_CSV = ROOT / "config/dimensions/task_type/canonical.csv"
ALIASES_CSV = ROOT / "config/dimensions/task_type/aliases_candidate.csv"
DATASET = RUN / "artifacts/unified_dataset.csv"
OUT = RUN / "artifacts/dimension_review/task_type_official_mapping_review_round3.csv"


def text(value: object) -> str:
    return value.isoformat() if isinstance(value, (datetime, date)) else str(value or "")


def normalized(value: str) -> str:
    return value.casefold().strip().replace(" ", "")


def main() -> None:
    wb = load_workbook(OFFICIAL_FILE, read_only=True, data_only=False)
    ws = wb[wb.sheetnames[0]]
    values = list(ws.iter_rows(values_only=True))
    wb.close()
    headers = list(values[0])
    id_idx, name_idx, numeric_idx = headers.index("数据唯一编号"), headers.index("订单类型"), headers.index("数值")
    canonicals = []
    for row in values[1:]:
        name = text(row[name_idx]).strip()
        if not name:
            continue
        numeric = row[numeric_idx]
        source_id = text(row[id_idx]).strip()
        canonicals.append({
            "official_task_type_id": text(numeric).strip() or source_id,
            "official_task_type_id_source": "数值" if numeric not in (None, "") else "数据唯一编号",
            "official_order_type": name,
            "official_source_record_id": source_id,
            "official_numeric_value": text(numeric).strip(),
        })
    if len({item['official_order_type'] for item in canonicals}) != len(canonicals):
        raise ValueError("official order type names must be unique")
    by_norm = {normalized(item["official_order_type"]): item for item in canonicals}
    by_name = {item["official_order_type"]: item for item in canonicals}
    with CANONICAL_CSV.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(canonicals[0])); writer.writeheader(); writer.writerows(canonicals)

    with DATASET.open(encoding="utf-8-sig", newline="") as handle:
        data = list(csv.DictReader(handle))
    inventory = defaultdict(lambda: {"count": 0, "source_ids": set(), "years": set(), "departments": set()})
    for row in data:
        raw = (row.get("task_type") or "").strip()
        stat = inventory[raw]; stat["count"] += 1; stat["source_ids"].add(row["source_id"]); stat["years"].add(row["year"]); stat["departments"].add(row["department"])

    # Read all source rows once per workbook to obtain source-only notes/context.
    source_rows = {}
    for record in data:
        raw = (record.get("task_type") or "").strip()
        source_rows.setdefault(record["source_file"], None)
    for filename in source_rows:
        raw_wb = load_workbook(RUN / "input" / filename, read_only=True, data_only=False)
        raw_ws = raw_wb[raw_wb.sheetnames[0]]
        raw_values = list(raw_ws.iter_rows(values_only=True))
        raw_wb.close()
        source_rows[filename] = (raw_values[0], raw_values)
    context = defaultdict(list)
    for row in data:
        raw = (row.get("task_type") or "").strip()
        headers, source_values = source_rows[row["source_file"]]
        original = {str(key): value for key, value in zip(headers, source_values[int(row["source_row_id"]) - 1])}
        original_task_field = "作业形式" if "作业形式" in original else ("作业类型" if "作业类型" in original else "咨询内容")
        course = original.get("专业/课程", original.get("专业", ""))
        notes = " | ".join(text(original.get(key)) for key in ("跟进反馈", "客户备注", "未成交原因") if original.get(key) not in (None, "", "/"))
        context[raw].append(f"{row['source_id']}:{row['source_row_id']} [原始{original_task_field}={text(original.get(original_task_field))}; 课程/专业={text(course) or '—'}; 备注={notes or '—'}]")

    def candidate(name: str, classification: str, confidence: str, evidence: str) -> tuple[str, str, str, str]:
        if name not in by_name:
            raise ValueError(f"non-official canonical attempted: {name}")
        return name, classification, confidence, evidence

    def classify(raw: str) -> tuple[str, str, str, str]:
        value = raw.strip(); norm = normalized(value)
        if not value or value == "/": return "", "UNKNOWN", "UNKNOWN", "空值或“/”未表达任务类型。"
        if any(token in value for token in ("+", "＋", "➕", "，")):
            return "", "MULTI_TASK", "MULTI_TASK", "包含多个任务语义；不得强制压缩到单一官方订单类型。"
        if norm in by_norm:
            item = by_norm[norm]
            return item["official_order_type"], "EXACT_MATCH", "HIGH", "与公司官方订单类型完全一致（忽略大小写与空格）。"
        if value in {"预存", "毕业无忧定金", "SVIP预存", "vip充值", "包课补款", "预存升级学年包定金"}:
            return "", "NON_TASK", "HIGH", "预存、充值、定金或补款为交易动作，不是公司官方订单类型。"
        if "补考" in value or "重写" in value:
            return "", "REVIEW_REQUIRED", "LOW", "官方范围区分“补考/重写-内部”与“补考/重写-外接”；原始记录未给出该归属。"
        if value in {"毕业论文", "毕业论文全包", "毕业论文半包"}:
            return candidate("Dissertation", "PROPOSED_HIGH", "HIGH", "中文“毕业论文”与官方 Dissertation 语义直接对应；服务包信息保留在原值。")
        if "毕业论文part" in norm or ("毕业论文" in value and "part" in norm):
            return candidate("Dissertation-part", "PROPOSED_HIGH", "HIGH", "明确包含毕业论文与 part，直接对应官方 Dissertation-part。")
        if "毕业论文" in value and "辅导" in value:
            return candidate("Dissertation- tutoring", "PROPOSED_HIGH", "HIGH", "明确包含毕业论文与辅导，直接对应官方 Dissertation- tutoring。")
        if value in {"论文润色", "润色", "浅润", "深润", "深度润色", "润色修改"} or "润色" in value:
            return candidate("润色-proofreading", "PROPOSED_HIGH", "HIGH", "包含明确润色服务语义，对应官方 润色-proofreading。")
        if value in {"论文", "期末论文", "15000词论文", "1W词论文"}:
            return candidate("Paper", "PROPOSED_MEDIUM", "MEDIUM", "可对应官方 Paper，但“论文”可能指学位论文或课程论文，需人工确认。")
        if "大论文" in value:
            return candidate("Dissertation", "PROPOSED_MEDIUM", "MEDIUM", "“大论文”通常指 Dissertation，但服务/组成部分语义不总是单一官方类型。")
        if "essay" in norm:
            return candidate("essay", "PROPOSED_HIGH", "HIGH", "包含明确 essay 语义；字数为描述信息。")
        if value in {"报告"}:
            return candidate("report", "PROPOSED_HIGH", "HIGH", "“报告”与官方 report 为直接中文对应。")
        if "report" in norm:
            return candidate("report", "PROPOSED_HIGH", "HIGH", "包含明确 report 语义。")
        if value in {"作业", "小组作业", "期末作业", "作业essay"}:
            return candidate("assignment", "PROPOSED_MEDIUM", "MEDIUM", "作业可能是 assignment、essay、report 等；官方 assignment 是候选而非已确认映射。")
        if "做题" in value or "试卷" in value:
            return candidate("做题", "PROPOSED_HIGH", "HIGH", "明确做题/试卷语义，对应官方 做题。")
        if "考试" in value or value in {"线下考", "线上考试"}:
            return candidate("考试", "PROPOSED_MEDIUM", "MEDIUM", "含考试语义，但与官方 online test-exam/quiz 的边界需人工确认。")
        if value in {"辅导", "考前辅导", "考试辅导"}:
            return candidate("tutoring", "PROPOSED_MEDIUM", "MEDIUM", "辅导语义可对应 tutoring，但官方还含提分/陪跑/包过/语言班等细分类型。")
        if value in {"学年包"}:
            return candidate("辅导年包", "PROPOSED_HIGH", "HIGH", "“学年包”与官方 辅导年包 为直接业务名称对应。")
        if value in {"半包课", "半包", "毕业无忧", "安心包", "卓越安心包", "svip", "SVIP", "VIP", "vip"}:
            return "", "REVIEW_REQUIRED", "LOW", "服务包名称未在官方订单类型中出现同名唯一项，不能自行选择陪跑课、包过辅导或辅导年包。"
        if value in {"简历", "简历制作"}:
            return candidate("CV/PS", "PROPOSED_HIGH", "HIGH", "简历制作可直接落入官方 CV/PS 范围。")
        if value in {"反思"}:
            return candidate("reflect", "PROPOSED_HIGH", "HIGH", "反思与官方 reflect 语义直接对应。")
        if value in {"ppt", "PPT", "PRE", "pre", "ppt演讲稿", "10分钟pre"}:
            return candidate("presentation", "PROPOSED_MEDIUM", "MEDIUM", "演示语义可对应 presentation，但部分记录可能是 Group Presentation 或仅演讲稿。")
        if value in {"小组pre"}:
            return candidate("Group Presentation", "PROPOSED_HIGH", "HIGH", "小组展示语义对应官方 Group Presentation。")
        if "海报" in value:
            return candidate("海报", "PROPOSED_HIGH", "HIGH", "包含明确海报交付物语义。")
        if "代码" in value or "matlab" in norm or "编程" in value:
            return candidate("code/experiment", "PROPOSED_HIGH", "HIGH", "包含代码/编程语义，对应官方 code/experiment。")
        if value in {"实验"}:
            return candidate("code/experiment", "PROPOSED_MEDIUM", "MEDIUM", "实验可能对应官方 code/experiment，但未说明是否包含代码。")
        if value in {"project"}:
            return candidate("project", "EXACT_MATCH", "HIGH", "与官方 project 完全一致。")
        if value in {"毕业设计"}:
            return candidate("project", "PROPOSED_MEDIUM", "MEDIUM", "毕业设计可为 project，但也可能为 Team Project，需确认。")
        if value in {"网课"}:
            return candidate("course work", "PROPOSED_MEDIUM", "MEDIUM", "网课可能对应 course work，但记录不足以确认服务性质。")
        if value in {"质检", "论文质检", "毕业论文质检", "大论文质检"}:
            return "", "REVIEW_REQUIRED", "LOW", "官方区分普通质检与高级质检；原始值未注明等级。"
        if value in {"quiz", "线上考试"}:
            return "", "REVIEW_REQUIRED", "LOW", "官方仅有组合类型 online test-exam/quiz；单独 quiz 不应自动合并。"
        if value in {"LR", "lr", "ME", "me", "文献综述", "文献综述部分", "ME部分", "毕业论文part", "大论文部分", "大论文me部分", "大论文LR和ME6000词"}:
            return "", "REVIEW_REQUIRED", "LOW", "论文部分缩写/文献综述无法在官方 Dissertation、Dissertation-part、bibliography、framework 等之间唯一选择。"
        return "", "REVIEW_REQUIRED", "LOW", "无充分证据将该原始值唯一映射到公司官方订单类型。"

    results = []; counts = Counter()
    for raw, stat in sorted(inventory.items(), key=lambda item: (-item[1]["count"], item[0])):
        proposed, classification, confidence, evidence = classify(raw)
        official = by_name.get(proposed, {})
        samples = " || ".join(context[raw][:3]) if context[raw] else "NOT_AVAILABLE"
        result = {"raw_value":raw,"record_count":stat["count"],"classification":classification,"proposed_official_task_type":proposed,"official_task_type_id":official.get("official_task_type_id", ""),"confidence":confidence,"evidence":evidence,"sample_context":samples,"final_decision":"","final_canonical":"","review_note":""}
        results.append(result); counts[classification] += 1
    fields = list(results[0])
    with OUT.open("w", encoding="utf-8", newline="") as handle:
        writer=csv.DictWriter(handle,fieldnames=fields); writer.writeheader(); writer.writerows(results)
    with ALIASES_CSV.open("w", encoding="utf-8", newline="") as handle:
        alias_fields=["raw_value","record_count","proposed_official_task_type","official_task_type_id","classification","confidence","evidence","source_ids","years","departments","approval_status"]
        writer=csv.DictWriter(handle,fieldnames=alias_fields); writer.writeheader()
        for result in results:
            stat=inventory[result["raw_value"]]
            writer.writerow({**{key:result[key] for key in ["raw_value","record_count","proposed_official_task_type","official_task_type_id","classification","confidence","evidence"]},"source_ids":"|".join(sorted(stat["source_ids"])),"years":"|".join(sorted(stat["years"])),"departments":"|".join(sorted(stat["departments"])),"approval_status":"CANDIDATE_ONLY"})
    if len(results) != 210 or sum(result["record_count"] for result in results) != 818:
        raise ValueError("incomplete task-type mapping inventory")
    print({"official_order_type_count":len(canonicals),"raw_unique_count":len(results),**counts})


if __name__ == "__main__": main()
