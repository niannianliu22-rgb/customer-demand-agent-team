#!/usr/bin/env python3
"""Prepare proposal-only Round-2 task-type review artifacts with raw evidence."""
from __future__ import annotations

import csv
from collections import defaultdict
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
RUN = ROOT / "runs/RUN-202608-DEMAND-001"
ARTIFACTS = RUN / "artifacts"
DIMENSION = ARTIFACTS / "dimension_review"
ROUND1 = DIMENSION / "task_type_unstandardized_review_round1.csv"
DATASET = ARTIFACTS / "unified_dataset.csv"


def stringify(value: object) -> str:
    return value.isoformat() if isinstance(value, (datetime, date)) else str(value)


def task_label(text: str) -> str:
    lower = text.lower()
    if "补考" in text: return "RESIT_EXAM"
    if "考试" in text or "考" in text: return "EXAM"
    if "essay" in lower: return "ESSAY"
    if "润色" in text or "浅润" in text or "深润" in text: return "PROOFREADING"
    if "续写" in text or "补写" in text or "重写" in text: return "CONTINUATION_OR_REWRITE"
    if "数据" in text: return "DATA_ANALYSIS_OR_COLLECTION"
    if "report" in lower or "报告" in text: return "REPORT"
    if "ppt" in lower or "pre" in lower or "演讲稿" in text: return "PRESENTATION"
    if "海报" in text: return "POSTER"
    if "视频" in text: return "VIDEO"
    if "选课" in text: return "COURSE_SELECTION"
    if "入学测试" in text: return "ADMISSION_TEST"
    if "quiz" in lower: return "QUIZ"
    if "作业" in text: return "ASSIGNMENT"
    if "毕业论文" in text or "大论文" in text: return "DISSERTATION"
    if "语言班" in text: return "LANGUAGE_COURSEWORK"
    return "UNSPECIFIED_TASK"


def multi_components(raw: str) -> list[str]:
    text = raw.replace("➕", "+").replace("＋", "+").replace("，", "+")
    parts = [part.strip() for part in text.split("+") if part.strip()]
    labels = []
    for part in parts:
        label = task_label(part)
        if label not in labels: labels.append(label)
    return labels or ["UNSPECIFIED_TASK"]


def main() -> None:
    with ROUND1.open(encoding="utf-8-sig", newline="") as handle:
        review_rows = list(csv.DictReader(handle))
    by_raw = {row["raw_value"]: row for row in review_rows}
    with DATASET.open(encoding="utf-8-sig", newline="") as handle:
        dataset_rows = list(csv.DictReader(handle))
    contexts = defaultdict(list)
    all_scope = {row["raw_value"] for row in review_rows if row["classification"] in {"PROPOSED_HIGH", "PROPOSED_MEDIUM", "REVIEW_REQUIRED", "UNKNOWN"}}
    workbook_cache = {}
    for record in dataset_rows:
        raw = (record.get("task_type") or "").strip()
        if raw not in all_scope:
            continue
        file_path = RUN / "input" / record["source_file"]
        if file_path not in workbook_cache:
            workbook = load_workbook(file_path, read_only=True, data_only=False)
            worksheet = workbook[record["source_sheet"]]
            headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            workbook_cache[file_path] = (workbook, worksheet, headers)
        _, worksheet, headers = workbook_cache[file_path]
        row_id = int(record["source_row_id"])
        values = [cell.value for cell in next(worksheet.iter_rows(min_row=row_id, max_row=row_id))]
        raw_record = {str(key): value for key, value in zip(headers, values)}
        task_field = "作业形式" if "作业形式" in raw_record else ("作业类型" if "作业类型" in raw_record else "咨询内容")
        course = raw_record.get("专业/课程", raw_record.get("专业", ""))
        notes = " | ".join(stringify(raw_record[field]) for field in ("跟进反馈", "客户备注", "未成交原因") if raw_record.get(field) not in (None, "", "/"))
        contexts[raw].append(
            f"{record['source_id']}:{record['source_row_id']} [原始{task_field}={raw_record.get(task_field, '')}; 课程/专业={course or '—'}; 学历={raw_record.get('学历', '—')}; 备注={notes or '—'}]"
        )
    for workbook, _, _ in workbook_cache.values(): workbook.close()

    def sample(raw: str) -> str:
        return " || ".join(contexts[raw][:3]) if contexts[raw] else "NOT_AVAILABLE"

    high = [row for row in review_rows if row["classification"] == "PROPOSED_HIGH"]
    manual = [row for row in review_rows if row["classification"] in {"PROPOSED_MEDIUM", "REVIEW_REQUIRED", "UNKNOWN"}]
    multi = [row for row in review_rows if row["is_multi_task"] == "MULTI_TASK"]
    non_task = [row for row in review_rows if row["classification"] == "NON_TASK"]

    high_fields = ["raw_value","record_count","proposed_canonical_task_type","proposed_task_type_group","confidence","reasoning","sample_context","audit_decision","final_canonical","final_group","review_note"]
    with (DIMENSION / "task_type_proposed_high_audit.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=high_fields); writer.writeheader()
        for row in high:
            writer.writerow({"raw_value":row["raw_value"],"record_count":row["record_count"],"proposed_canonical_task_type":row["proposed_canonical_task_type"],"proposed_task_type_group":row["proposed_task_type_group"],"confidence":row["confidence"],"reasoning":row["reasoning"],"sample_context":sample(row["raw_value"]),"audit_decision":"","final_canonical":"","final_group":"","review_note":""})
    manual_fields = ["raw_value","record_count","classification","proposed_canonical_task_type","proposed_task_type_group","confidence","reasoning","sample_context","decision","final_canonical","final_group","review_note"]
    with (DIMENSION / "task_type_manual_review_round2.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=manual_fields); writer.writeheader()
        for row in manual:
            writer.writerow({"raw_value":row["raw_value"],"record_count":row["record_count"],"classification":row["classification"],"proposed_canonical_task_type":row["proposed_canonical_task_type"],"proposed_task_type_group":row["proposed_task_type_group"],"confidence":row["confidence"],"reasoning":row["reasoning"],"sample_context":sample(row["raw_value"]),"decision":"","final_canonical":"","final_group":"","review_note":""})
    multi_fields = ["raw_value","record_count","detected_tasks","proposed_task_1","proposed_task_2","proposed_task_3","recommended_handling","reasoning"]
    with (DIMENSION / "task_type_multi_task_review.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=multi_fields); writer.writeheader()
        for row in multi:
            tasks = multi_components(row["raw_value"])
            writer.writerow({"raw_value":row["raw_value"],"record_count":row["record_count"],"detected_tasks":"|".join(tasks),"proposed_task_1":tasks[0] if tasks else "","proposed_task_2":tasks[1] if len(tasks)>1 else "","proposed_task_3":tasks[2] if len(tasks)>2 else "","recommended_handling":"Preserve primary_task_type plus task_type_list; require human confirmation of primary task.","reasoning":f"{row['reasoning']} 原始 Excel 样本：{sample(row['raw_value'])}"})
    other_field = {
        "毕业无忧定金": "payment_stage / deposit_type",
        "SVIP预存": "payment_transaction_type / prepaid_credit",
        "vip充值": "payment_transaction_type / top_up",
        "包课补款": "payment_transaction_type / balance_payment",
    }
    non_fields = ["raw_value","record_count","classification","reasoning","likely_field_or_data_type","sample_context","review_note"]
    with (DIMENSION / "task_type_non_task_review.csv").open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=non_fields); writer.writeheader()
        for row in non_task:
            writer.writerow({"raw_value":row["raw_value"],"record_count":row["record_count"],"classification":"NON_TASK","reasoning":row["reasoning"],"likely_field_or_data_type":other_field.get(row["raw_value"],"commercial_transaction"),"sample_context":sample(row["raw_value"]),"review_note":""})
    if len(high) != 95 or len(manual) != 85 or len(multi) != 20 or len(non_task) != 4:
        raise ValueError("Round-1 classifications unexpectedly changed")


if __name__ == "__main__": main()
