#!/usr/bin/env python3
"""Deep-trace selected ambiguous school values against raw Excel rows.

This is an evidence artifact only.  It does not amend the dataset or aliases.
"""
from __future__ import annotations

import csv
import json
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
RUN = ROOT / "runs/RUN-202608-DEMAND-001"
DATASET = RUN / "artifacts/unified_dataset.csv"
OUTPUT = RUN / "artifacts/dimension_review/school_ambiguous_deep_trace_round3.csv"

TARGETS = {"波士顿", "维多利亚", "UCD", "加州大学", "北卡", "华盛顿", "多伦多高中", "麦唐纳国际学校"}
ASSESSMENTS = {
    "波士顿": ("", "UNRESOLVED", "UNRESOLVED", "三条原始记录均只出现城市名；跟进信息仅描述作业/成交，未包含院校、学院、课程代码或邮箱域名。"),
    "维多利亚": ("", "UNRESOLVED", "UNRESOLVED", "两条原始记录分别写英国和澳洲，备注均无学校身份信息；存在潜在多实体风险，但尚无证据确认实际为不同学校。"),
    "UCD": ("", "UNRESOLVED", "UNRESOLVED", "原始咨询仅写 UCD 与选课，未提供课程、校区、学院或备注证据。"),
    "加州大学": ("", "UNRESOLVED", "UNRESOLVED", "原始跟进反馈仅说明写作课包课，未提供 UC 系统的具体校区。"),
    "北卡": ("", "UNRESOLVED", "UNRESOLVED", "原始未成交原因说明网站制作报价，但没有 UNC 的具体校区或正式英文名称。"),
    "华盛顿": ("", "UNRESOLVED", "UNRESOLVED", "原始跟进反馈仅写金融考试，不能在 University of Washington 与 Washington University in St. Louis 等之间消歧。"),
    "多伦多高中": ("", "NON_UNIVERSITY_ENTITY", "NOT_APPLICABLE", "原始学历为高中，原始学校值为“多伦多高中”；未提供正式英文名称或地址，不能识别为具体中学。"),
    "麦唐纳国际学校": ("", "NON_UNIVERSITY_ENTITY", "NOT_APPLICABLE", "原始学历为高中、国家为加拿大；为国际学校实体而非大学，原始行未提供可唯一识别的英文名称或地址。"),
}
MANUAL_CANDIDATES = {
    "cmu": ("Carnegie Mellon University", "CANDIDATE_PENDING_HUMAN", "美国硕士、计算机基础与原始“2h考试”提供弱辅助上下文；缩写本身仍可能对应其他学校。"),
    "psu": ("Pennsylvania State University", "CANDIDATE_PENDING_HUMAN", "美国本科心理学提供弱辅助上下文；PSU 仍可能对应其他学校。"),
    "伯克利": ("University of California, Berkeley", "CANDIDATE_PENDING_HUMAN", "原始跟进反馈写“美国伯克利选课和入学测试题”，显著支持 UC Berkeley 候选；仍保留人工确认要求。"),
}


def json_default(value: object) -> str:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return str(value)


def main() -> None:
    with DATASET.open(encoding="utf-8-sig", newline="") as handle:
        records = [row for row in csv.DictReader(handle) if (row.get("school") or "").strip() in TARGETS | set(MANUAL_CANDIDATES)]
    expected = TARGETS | set(MANUAL_CANDIDATES)
    if {row["school"] for row in records} != expected:
        raise ValueError("the target records in unified_dataset do not match the review scope")

    workbook_cache = {}
    rows = []
    for record in records:
        file_path = RUN / "input" / record["source_file"]
        if file_path not in workbook_cache:
            workbook = load_workbook(file_path, read_only=True, data_only=False)
            worksheet = workbook[record["source_sheet"]]
            headers = [cell.value for cell in next(worksheet.iter_rows(min_row=1, max_row=1))]
            workbook_cache[file_path] = (workbook, worksheet, headers)
        _, worksheet, headers = workbook_cache[file_path]
        source_row = int(record["source_row_id"])
        raw_values = [cell.value for cell in next(worksheet.iter_rows(min_row=source_row, max_row=source_row))]
        raw = {str(key): value for key, value in zip(headers, raw_values)}
        raw_school_field = "学校" if "学校" in raw else "院校"
        raw_notes = " | ".join(str(raw.get(field)) for field in ("跟进反馈", "客户备注", "未成交原因") if raw.get(field) not in (None, "", "/"))
        raw_course = raw.get("专业/课程", raw.get("专业", ""))
        raw_task = raw.get("作业形式", raw.get("咨询内容", raw.get("作业类型", "")))
        raw_order_id = raw.get("订单编号", "")
        value = record["school"]
        if value in MANUAL_CANDIDATES:
            recommended, confidence, evidence = MANUAL_CANDIDATES[value]
            resolution_status = "MANUAL_CONFIRMATION_CANDIDATE"
            alias_status = "NO_CONFIRMED_MULTI_ENTITY_ALIAS"
        else:
            recommended, confidence, resolution_status, evidence = ASSESSMENTS[value]
            alias_status = "POTENTIAL_MULTI_ENTITY_ALIAS" if value == "维多利亚" else "NO_CONFIRMED_MULTI_ENTITY_ALIAS"
        rows.append({
            "raw_value": value,
            "source_id": record["source_id"],
            "source_file": record["source_file"],
            "source_sheet": record["source_sheet"],
            "source_row_id": record["source_row_id"],
            "record_id": raw_order_id or record.get("order_id", "") or "NOT_AVAILABLE",
            "raw_school_field": raw_school_field,
            "raw_school_value": raw.get(raw_school_field, ""),
            "raw_country": raw.get("国家", ""),
            "raw_degree": raw.get("学历", ""),
            "raw_major_or_course": raw_course,
            "raw_task_or_consultation": raw_task,
            "raw_consultation_date": raw.get("日期", ""),
            "raw_ddl": raw.get("DDL", ""),
            "raw_notes": raw_notes or "NOT_AVAILABLE",
            "course_code": "NOT_AVAILABLE_IN_SOURCE_ROW",
            "email_domain": "NOT_AVAILABLE_IN_SOURCE_ROW",
            "resolved_canonical": "",
            "recommended_canonical": recommended,
            "resolution_status": resolution_status,
            "confidence": confidence,
            "evidence": evidence,
            "multi_entity_alias_status": alias_status,
            "raw_record_json": json.dumps(raw, ensure_ascii=False, default=json_default, sort_keys=True),
        })
    for workbook, _, _ in workbook_cache.values():
        workbook.close()
    fields = list(rows[0])
    with OUTPUT.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)
    if len(rows) != 14:
        raise ValueError(f"expected 14 scoped records, got {len(rows)}")


if __name__ == "__main__":
    main()
