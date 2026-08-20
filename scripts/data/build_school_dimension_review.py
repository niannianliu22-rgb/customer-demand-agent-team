"""Build a read-only school-value inventory and non-binding alias candidates."""

from __future__ import annotations

import csv
import json
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]

# Curated entity hypotheses.  These only create PROPOSED/REVIEW_REQUIRED
# candidates; no group is allowed to alter unified_dataset in this script.
CANDIDATES = [
    ("SCH-CAND-001", ["新南", "新南威尔士", "unsw", "UNSW"], "University of New South Wales", "Australia", "HIGH", "PROPOSED"),
    ("SCH-CAND-002", ["ucl", "UCL"], "University College London", "United Kingdom", "HIGH", "PROPOSED"),
    ("SCH-CAND-003", ["kcl", "KCL"], "King's College London", "United Kingdom", "HIGH", "PROPOSED"),
    ("SCH-CAND-004", ["悉大", "悉尼大学"], "University of Sydney", "Australia", "HIGH", "PROPOSED"),
    ("SCH-CAND-005", ["曼大", "曼彻斯特"], "University of Manchester", "United Kingdom", "MEDIUM", "PROPOSED"),
    ("SCH-CAND-006", ["南安", "南安普顿"], "University of Southampton", "United Kingdom", "HIGH", "PROPOSED"),
    ("SCH-CAND-007", ["墨尔本", "墨大"], "University of Melbourne", "Australia", "MEDIUM", "PROPOSED"),
    ("SCH-CAND-008", ["格拉", "格拉斯哥"], "University of Glasgow", "United Kingdom", "MEDIUM", "PROPOSED"),
    ("SCH-CAND-009", ["昆士兰", "昆士兰大学", "UQ"], "University of Queensland", "Australia", "HIGH", "PROPOSED"),
    ("SCH-CAND-010", ["谢菲", "谢菲尔德"], "University of Sheffield", "United Kingdom", "HIGH", "PROPOSED"),
    ("SCH-CAND-011", ["华威", "华威大学"], "University of Warwick", "United Kingdom", "HIGH", "PROPOSED"),
    ("SCH-CAND-012", ["阿德莱德", "阿德"], "University of Adelaide", "Australia", "HIGH", "PROPOSED"),
    ("SCH-CAND-013", ["澳国立", "澳洲国立大学"], "Australian National University", "Australia", "HIGH", "PROPOSED"),
    ("SCH-CAND-014", ["香港大学", "港大"], "The University of Hong Kong", "Hong Kong", "HIGH", "PROPOSED"),
    ("SCH-CAND-015", ["香港理工", "香港理工大学"], "The Hong Kong Polytechnic University", "Hong Kong", "HIGH", "PROPOSED"),
    ("SCH-CAND-016", ["香港城市大学", "港城", "港城大"], "City University of Hong Kong", "Hong Kong", "MEDIUM", "PROPOSED"),
    ("SCH-CAND-017", ["香港教育大学", "港教"], "The Education University of Hong Kong", "Hong Kong", "HIGH", "PROPOSED"),
    ("SCH-CAND-018", ["纽约大学", "NYU"], "New York University", "United States", "HIGH", "PROPOSED"),
    ("SCH-CAND-019", ["奥克兰", "奥克兰大学"], "University of Auckland", "New Zealand", "MEDIUM", "PROPOSED"),
    ("SCH-CAND-020", ["ual"], "University of the Arts London", "United Kingdom", "REVIEW_REQUIRED", "REVIEW_REQUIRED"),
    ("SCH-CAND-021", ["qm"], "Queen Mary University of London", "United Kingdom", "REVIEW_REQUIRED", "REVIEW_REQUIRED"),
    ("SCH-CAND-022", ["国立"], None, None, "REVIEW_REQUIRED", "REVIEW_REQUIRED"),
    ("SCH-CAND-023", ["伦敦"], None, None, "REVIEW_REQUIRED", "REVIEW_REQUIRED"),
    ("SCH-CAND-024", ["约克"], None, None, "REVIEW_REQUIRED", "REVIEW_REQUIRED"),
    ("SCH-CAND-025", ["纽卡"], None, None, "REVIEW_REQUIRED", "REVIEW_REQUIRED"),
]


def normalize(value: str) -> str:
    return re.sub(r"\s+", " ", unicodedata.normalize("NFKC", value).strip()).casefold()


def main(run_id: str):
    artifacts = ROOT / "runs" / run_id / "artifacts"
    output_dir = artifacts / "dimension_review"
    output_dir.mkdir(exist_ok=True)
    with (artifacts / "unified_dataset.csv").open(encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    required_trace = ["source_id", "source_file", "source_sheet", "source_row_id", "year", "department"]
    if any(any(not row.get(field) for field in required_trace) for row in rows):
        raise RuntimeError("unified dataset lacks complete traceability")

    # Re-read all original values by source-row pointer; never modify them.
    workbooks = {}
    raw_school_by_pointer = {}
    for row in rows:
        source_id = row["source_id"]
        if source_id in workbooks:
            continue
        path = ROOT / "runs" / run_id / "input" / row["source_file"]
        wb = load_workbook(path, read_only=True, data_only=True)
        workbooks[source_id] = wb
        for ws in wb.worksheets:
            headers = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
            field = "学校" if "学校" in headers else "院校" if "院校" in headers else None
            if field:
                index = headers.index(field)
                for excel_row, values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    raw_school_by_pointer[(source_id, ws.title, excel_row)] = values[index]

    inventory = {}
    trace_mismatches = []
    for row in rows:
        value = row["school"]
        pointer = (row["source_id"], row["source_sheet"], int(row["source_row_id"]))
        raw_value = raw_school_by_pointer.get(pointer)
        if (raw_value or "") != value:
            trace_mismatches.append({"pointer": pointer, "unified_value": value, "raw_value": raw_value})
        if not value.strip():
            continue
        entry = inventory.setdefault(value, {
            "original_value": value, "normalized_text": normalize(value), "count": 0,
            "source_ids": set(), "years": set(), "departments": set(), "country_values": set(), "sample_source_rows": []
        })
        entry["count"] += 1; entry["source_ids"].add(row["source_id"]); entry["years"].add(int(row["year"])); entry["departments"].add(row["department"])
        if row["country"].strip(): entry["country_values"].add(row["country"])
        if len(entry["sample_source_rows"]) < 5:
            entry["sample_source_rows"].append({"source_id": row["source_id"], "source_sheet": row["source_sheet"], "source_row_id": int(row["source_row_id"])})
    for wb in workbooks.values(): wb.close()

    values = []
    for entry in sorted(inventory.values(), key=lambda item: (-item["count"], item["original_value"])):
        for key in ("source_ids", "years", "departments", "country_values"):
            entry[key] = sorted(entry[key])
        values.append(entry)
    lookup = {entry["original_value"]: entry for entry in values}
    groups = []
    for group_id, variants, canonical, country, confidence, status in CANDIDATES:
        present = [lookup[value] for value in variants if value in lookup]
        if not present: continue
        groups.append({
            "candidate_group_id": group_id,
            "original_values": [entry["original_value"] for entry in present],
            "suggested_canonical_name": canonical,
            "suggested_country": country,
            "confidence": confidence,
            "evidence": {
                "variant_counts": {entry["original_value"]: entry["count"] for entry in present},
                "country_values": sorted({country for entry in present for country in entry["country_values"]}),
                "source_ids": sorted({source for entry in present for source in entry["source_ids"]}),
                "sample_source_rows": [sample for entry in present for sample in entry["sample_source_rows"]][:10],
                "basis": "Candidate only: observed spelling/case/abbreviation variants plus country and cross-source occurrence. No values were merged."
            },
            "status": status
        })
    inventory_doc = {
        "run_id": run_id, "artifact": "school_value_inventory", "source_dataset": "unified_dataset.csv",
        "business_row_count": len(rows), "nonempty_school_row_count": sum(entry["count"] for entry in values),
        "nonempty_unique_value_count": len(values), "traceability_validation": {"checked_rows": len(rows), "raw_source_mismatch_count": len(trace_mismatches), "mismatches": trace_mismatches}, "values": values
    }
    candidates_doc = {
        "run_id": run_id, "artifact": "school_alias_candidates", "source_inventory": "school_value_inventory.json",
        "candidate_group_count": len(groups), "confidence_counts": dict(sorted(defaultdict(int, {key: sum(1 for group in groups if group["confidence"] == key) for key in ("HIGH", "MEDIUM", "REVIEW_REQUIRED")}).items())),
        "groups": groups,
        "guardrail": "All groups are candidates only. No group is APPROVED and no school value has been replaced in unified_dataset."
    }
    (output_dir / "school_value_inventory.json").write_text(json.dumps(inventory_doc, ensure_ascii=False, indent=2), encoding="utf-8")
    (output_dir / "school_alias_candidates.json").write_text(json.dumps(candidates_doc, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({"unique_nonempty_values": len(values), "candidate_groups": len(groups), "confidence_counts": candidates_doc["confidence_counts"], "traceability_mismatches": len(trace_mismatches)}, ensure_ascii=False))


if __name__ == "__main__":
    main(sys.argv[1] if len(sys.argv) > 1 else "RUN-202608-DEMAND-001")
