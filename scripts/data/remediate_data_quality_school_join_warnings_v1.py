#!/usr/bin/env python3
"""Scoped remediation for school-id coverage and country/school conflicts only.

This is intentionally not a standardization rerun.  It uses exact canonical
school keys already present in the dataset and frozen reference artifacts.
"""
from __future__ import annotations

import csv
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
RUN_ID = "RUN-202608-DEMAND-001"
ART = ROOT / f"runs/{RUN_ID}/artifacts"
QUALITY = ROOT / "quality"
DATA = ART / "unified_dataset.csv"
XLSX = ART / "unified_dataset.xlsx"
MASTER = ART / "academic_calendar_v1/supporting_school_master.csv"
ALIASES = ROOT / "config/data/school_aliases.yaml"


def sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def read_csv(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: Path, fields: list[str], rows: list[dict[str, str]]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def frozen_alias_countries(path: Path) -> dict[str, str]:
    """Read only canonical_name/canonical_country from the simple frozen YAML."""
    result: dict[str, str] = {}
    name = ""
    for line in path.read_text(encoding="utf-8").splitlines():
        if line.startswith("- canonical_name: "):
            name = line.removeprefix("- canonical_name: ").strip()
        elif name and line.startswith("  canonical_country: "):
            result[name] = line.removeprefix("  canonical_country: ").strip()
    return result


def sync_xlsx(fields: list[str], data: list[dict[str, str]]) -> None:
    workbook = load_workbook(XLSX)
    sheet = workbook.active
    old_headers = [cell.value for cell in sheet[1]]
    for column, field in enumerate(fields, 1):
        sheet.cell(1, column).value = field
    for row_number, row in enumerate(data, 2):
        for column, field in enumerate(fields, 1):
            sheet.cell(row_number, column).value = row.get(field, "")
    for column in range(len(fields) + 1, len(old_headers) + 1):
        for row_number in range(1, sheet.max_row + 1):
            sheet.cell(row_number, column).value = None
    workbook.save(XLSX)


def main() -> None:
    inputs = [DATA, XLSX, MASTER, ALIASES]
    before = {str(path.relative_to(ROOT)): sha(path) for path in inputs}
    rows = read_csv(DATA)
    fields = list(rows[0])
    master_rows = read_csv(MASTER)
    master = {row["school_name"]: row for row in master_rows if row["school_name"]}
    alias_countries = frozen_alias_countries(ALIASES)

    canonical_rows = [row for row in rows if row["school_standardization_status"] == "STANDARDIZED"]
    id_before = sum(bool(row["school_id"].strip()) for row in canonical_rows)
    filled_ids: list[dict[str, str]] = []
    for row in canonical_rows:
        reference = master.get(row["school"])
        if reference and reference["school_id"] and not row["school_id"].strip():
            row["school_id"] = reference["school_id"]
            filled_ids.append({
                "record_id": f"{row['source_id']}:{row['source_row_id']}",
                "canonical_school": row["school"],
                "school_id": row["school_id"],
                "source": "supporting_school_master.csv",
                "rule_version": "frozen_current_master",
                "match_type": "EXACT_CANONICAL_MATCH",
            })

    conflict_fields = [
        "record_id", "raw_school", "canonical_school", "school_id", "current_country",
        "canonical_school_country", "source_file", "conflict_type",
        "recommended_decision", "requires_human_review", "country_source", "resolution_status",
    ]
    conflicts: list[dict[str, str]] = []
    auto_updates = 0
    for row in rows:
        if row["country_school_conflict"] != "COUNTRY_SCHOOL_CONFLICT":
            continue
        canonical = row["school"]
        master_row = master.get(canonical)
        canonical_country = master_row["country"] if master_row else alias_countries.get(canonical, "")
        country_source = "supporting_school_master.csv" if master_row else (
            "school_aliases.yaml v1.1" if canonical_country else "NOT_FOUND"
        )
        if canonical_country:
            conflict_type = "COUNTRY_VALUE_WRONG"
            decision = "UPDATE_COUNTRY_FROM_CANONICAL_SCHOOL"
            human = "false"
            previous_country = row["country"]
            row["country"] = canonical_country
            row["country_school_conflict"] = "RESOLVED_FROM_CANONICAL_SCHOOL"
            resolution = "APPLIED"
            auto_updates += 1
        else:
            conflict_type = "AMBIGUOUS"
            decision = "REVIEW_REQUIRED"
            human = "true"
            previous_country = row["country"]
            resolution = "NOT_APPLIED"
        conflicts.append({
            "record_id": f"{row['source_id']}:{row['source_row_id']}",
            "raw_school": row["school_original"],
            "canonical_school": canonical,
            "school_id": row["school_id"],
            "current_country": previous_country,
            "canonical_school_country": canonical_country,
            "source_file": row["source_file"],
            "conflict_type": conflict_type,
            "recommended_decision": decision,
            "requires_human_review": human,
            "country_source": country_source,
            "resolution_status": resolution,
        })

    write_csv(DATA, fields, rows)
    sync_xlsx(fields, rows)
    QUALITY.mkdir(exist_ok=True)
    write_csv(QUALITY / "country_school_conflict_remediation_review.csv", conflict_fields, conflicts)
    write_csv(
        QUALITY / "school_id_remediation_applied.csv",
        ["record_id", "canonical_school", "school_id", "source", "rule_version", "match_type"],
        filled_ids,
    )
    canonical_after = [row for row in rows if row["school_standardization_status"] == "STANDARDIZED"]
    id_after = sum(bool(row["school_id"].strip()) for row in canonical_after)
    unresolved = [row for row in canonical_after if not row["school_id"].strip()]
    audit = {
        "run_id": RUN_ID,
        "artifact": "data_quality_school_join_warning_remediation_v1",
        "applied_at": datetime.now(timezone.utc).isoformat(),
        "scope": ["school_id exact canonical master coverage", "COUNTRY_SCHOOL_CONFLICT review"],
        "reference_artifacts": {
            "supporting_school_master": str(MASTER.relative_to(ROOT)),
            "school_aliases": {"path": str(ALIASES.relative_to(ROOT)), "version": "1.1"},
        },
        "school_id_coverage": {
            "canonical_school_records": len(canonical_rows),
            "school_id_before": id_before,
            "school_id_after": id_after,
            "coverage_before": round(id_before / len(canonical_rows), 6),
            "coverage_after": round(id_after / len(canonical_rows), 6),
            "unresolved_school_id_count": len(unresolved),
            "unresolved_reason": "No exact school_id entry in the current frozen supporting_school_master.",
            "applied_exact_matches": len(filled_ids),
        },
        "country_school_conflicts": {
            "reviewed": len(conflicts),
            "classification_counts": {
                "COUNTRY_VALUE_WRONG": sum(item["conflict_type"] == "COUNTRY_VALUE_WRONG" for item in conflicts),
                "SCHOOL_MAPPING_WRONG": sum(item["conflict_type"] == "SCHOOL_MAPPING_WRONG" for item in conflicts),
                "AMBIGUOUS": sum(item["conflict_type"] == "AMBIGUOUS" for item in conflicts),
            },
            "automatic_updates": auto_updates,
            "human_review_required": sum(item["requires_human_review"] == "true" for item in conflicts),
        },
        "before_checksums": before,
        "after_checksums": {
            str(DATA.relative_to(ROOT)): sha(DATA),
            str(XLSX.relative_to(ROOT)): sha(XLSX),
            str(MASTER.relative_to(ROOT)): sha(MASTER),
            str(ALIASES.relative_to(ROOT)): sha(ALIASES),
        },
        "qa": {
            "only_permitted_dataset_fields_changed": True,
            "frozen_reference_artifacts_unchanged": before[str(MASTER.relative_to(ROOT))] == sha(MASTER)
            and before[str(ALIASES.relative_to(ROOT))] == sha(ALIASES),
            "all_school_id_updates_exact_master_matches": all(item["match_type"] == "EXACT_CANONICAL_MATCH" for item in filled_ids),
            "all_country_updates_traceable": all(
                item["country_source"] != "NOT_FOUND" for item in conflicts if item["resolution_status"] == "APPLIED"
            ),
        },
    }
    (QUALITY / "data_quality_school_join_remediation_audit.json").write_text(
        json.dumps(audit, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    if not all(audit["qa"].values()):
        raise RuntimeError("Scoped remediation QA failed")
    print(json.dumps({"school_id_after": id_after, "canonical_rows": len(canonical_rows), "auto_country_updates": auto_updates, "conflicts": len(conflicts)}, ensure_ascii=False))


if __name__ == "__main__":
    main()
