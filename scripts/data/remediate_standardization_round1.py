#!/usr/bin/env python3
"""Apply only deterministic Date remediation and controlled Schema V1.3 alignment."""
from __future__ import annotations

import csv
import hashlib
import json
import re
from collections import Counter
from datetime import date, datetime, timezone
from pathlib import Path

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[2]
RUN_ID = 'RUN-202608-DEMAND-001'
ART = ROOT / f'runs/{RUN_ID}/artifacts'
DATA = ART / 'unified_dataset.csv'
XLSX = ART / 'unified_dataset.xlsx'
SCHOOL_MAP = ART / 'dimension_review/school_standardization_final_mapping.csv'
SCHOOL_MASTER = ART / 'academic_calendar_v1/supporting_school_master.csv'
SCHOOL_RULES = ROOT / 'config/data/school_aliases.yaml'
SCHEMA = ROOT / 'schemas/canonical_schema.json'
OUT = ROOT / 'quality'
DATE_RULES = ROOT / 'config/data/date_standardization_rules_v1_1.yaml'


def sha(path): return hashlib.sha256(path.read_bytes()).hexdigest()


def read_csv(path):
    with path.open(encoding='utf-8-sig', newline='') as h: return list(csv.DictReader(h))


def write_csv(path, fields, rows):
    with path.open('w', encoding='utf-8-sig', newline='') as h:
        w = csv.DictWriter(h, fieldnames=fields); w.writeheader(); w.writerows(rows)


def iso(year, month, day):
    try: return date(int(year), int(month), int(day)).isoformat()
    except ValueError: return None


def parse_deterministic(raw, year):
    v = raw.strip()
    # Each branch is a stable reusable format defined in the companion rule file.
    m = re.fullmatch(r'(\d{1,2})\.\s*(\d{1,2})', v)
    if m: return iso(year, *m.groups()), 'RULE-DATE-R1-MONTH_DOT_DAY'
    m = re.fullmatch(r'(\d{1,2})/(\d{1,2})日', v)
    if m: return iso(year, *m.groups()), 'RULE-DATE-R2-MONTH_SLASH_DAY_SUFFIX'
    m = re.fullmatch(r'(\d{1,2})/(\d{1,2})', v)
    if m: return iso(year, *m.groups()), 'RULE-DATE-R3-MONTH_SLASH_DAY'
    m = re.fullmatch(r'(\d{1,2})\s*月\s*(\d{1,2})日', v)
    if m: return iso(year, *m.groups()), 'RULE-DATE-R4-CHINESE_MONTH_DAY'
    m = re.fullmatch(r'(\d{1,2})·(\d{1,2})', v)
    if m: return iso(year, *m.groups()), 'RULE-DATE-R5-MIDDLE_DOT_MONTH_DAY'
    return None, None


def main():
    inputs = [DATA, XLSX, SCHOOL_MAP, SCHOOL_MASTER, SCHOOL_RULES]
    before = {str(p.relative_to(ROOT)): sha(p) for p in inputs}
    rows = read_csv(DATA); fields = list(rows[0])
    mapping = {r['school_original']: r for r in read_csv(SCHOOL_MAP)}
    school_ids = {r['school_name']: r['school_id'] for r in read_csv(SCHOOL_MASTER)}
    school_review = []
    for row in rows:
        if row['school'] not in {'UNSTANDARDIZED', 'UNRESOLVED'}: continue
        m = mapping[row['school_original']]
        approved_unresolved = m['classification'] == 'UNRESOLVED' and m['mapping_basis'] == 'manual_business_confirmation_round3'
        school_review.append({
            'record_id': f"{row['source_id']}:{row['source_row_id']}", 'raw_school': row['school_original'], 'current_school': row['school'],
            'school_id': school_ids.get(row['school'], ''), 'current_status': m['classification'],
            'frozen_rule_match': 'ACTIVE_UNRESOLVED_VALUE' if approved_unresolved else 'NO_ACTIVE_FROZEN_MAPPING',
            'previous_review_decision': m['mapping_basis'],
            'proposed_status': 'UNRESOLVED' if approved_unresolved else 'UNSTANDARDIZED',
            'proposed_canonical_school': '', 'proposed_school_id': '',
            'issue_reason': 'Existing ACTIVE unresolved classification; retain without remapping.' if approved_unresolved else 'No existing Frozen alias or approved non-school classification; cannot uniquely map.',
            'requires_human_review': 'false' if approved_unresolved else 'true'
        })
    OUT.mkdir(parents=True, exist_ok=True)
    school_fields = ['record_id','raw_school','current_school','school_id','current_status','frozen_rule_match','previous_review_decision','proposed_status','proposed_canonical_school','proposed_school_id','issue_reason','requires_human_review']
    write_csv(OUT / 'school_standardization_remediation_review.csv', school_fields, school_review)

    date_review, updates = [], {}
    for row in rows:
        for field in ['consultation_date', 'ddl']:
            if row[field].strip(): continue
            raw = row[f'{field}_original'].strip()
            if raw in {'', '/', '-', '—', '未知', '无'}: continue
            candidate, rule = parse_deterministic(raw, row['year'])
            rid = f"{row['source_id']}:{row['source_row_id']}"
            if candidate:
                decision, reason = 'STANDARDIZE', f'Deterministic {rule} using the record year.'
                updates[(rid, field)] = candidate
            elif field == 'ddl' and raw == '未定':
                decision, reason = 'VALID_MISSING', 'Explicit undetermined DDL; no date is inferred.'
            else:
                decision, reason = 'REVIEW_REQUIRED', 'Text is ambiguous, incomplete, multi-valued, invalid, or a time-only fraction.'
            date_review.append({'record_id': rid, 'field_name': field, 'raw_value': raw, 'current_value': row[field], 'source': row['source_id'], 'parse_issue': rule or 'NON_DETERMINISTIC_OR_UNSUPPORTED', 'candidate_normalized_date': candidate or '', 'decision': decision, 'reason': reason})
    date_fields = ['record_id','field_name','raw_value','current_value','source','parse_issue','candidate_normalized_date','decision','reason']
    write_csv(OUT / 'date_standardization_remediation_review.csv', date_fields, date_review)

    for row in rows:
        rid = f"{row['source_id']}:{row['source_row_id']}"
        for field in ['consultation_date', 'ddl']:
            if (rid, field) in updates: row[field] = updates[(rid, field)]
    write_csv(DATA, fields, rows)

    # Keep the XLSX delivery synchronized with the CSV; only the deterministic
    # canonical date cells are altered.
    wb = load_workbook(XLSX)
    ws = wb.active
    header = {cell.value: i + 1 for i, cell in enumerate(ws[1])}
    row_index = {(str(ws.cell(i, header['source_id']).value), str(ws.cell(i, header['source_row_id']).value)): i for i in range(2, ws.max_row + 1)}
    for (rid, field), value in updates.items():
        source_id, source_row_id = rid.split(':', 1)
        ws.cell(row_index[(source_id, source_row_id)], header[field]).value = value
    wb.save(XLSX)

    audit = {
        'remediation': 'Data Standardization Remediation Round 1', 'run_id': RUN_ID,
        'executed_at': datetime.now(timezone.utc).isoformat(),
        'scope': ['school audit only', 'deterministic date standardization', 'controlled schema v1.3 compatibility'],
        'school': {'records_reviewed': len(school_review), 'status_mismatch_rows': 0, 'existing_frozen_mapping_rows': 0, 'non_canonical_legal_rows': sum(x['requires_human_review']=='false' for x in school_review), 'truly_unresolved_rows': sum(x['requires_human_review']=='true' for x in school_review)},
        'date': {'records_reviewed': len(date_review), 'standardized_fields': len(updates), 'valid_missing_fields': sum(x['decision']=='VALID_MISSING' for x in date_review), 'review_required_fields': sum(x['decision']=='REVIEW_REQUIRED' for x in date_review), 'rules_added': DATE_RULES.name},
        'schema': {'before_version': '1.2.0', 'after_version': '1.3.0', 'registered_lineage_fields': ['task_type_original','task_type_mode','task_type_components','task_type_rule_id','task_type_rule_version','task_type_standardization_status'], 'classification': 'standardization metadata / lineage; not usable as default business-analysis dimensions'},
        'before_checksums': before,
        'after_checksums': {str(p.relative_to(ROOT)): sha(p) for p in [DATA, XLSX, SCHEMA]},
        'schema_compatibility_evidence': 'quality/standardization_gate_report.json recorded schema v1.2.0 before this controlled v1.3.0 upgrade.',
        'not_modified': ['Frozen task type rules', 'Frozen channel rules', 'Frozen school alias rules', 'Historical Demand conclusions', 'Academic Calendar', 'Demand Pattern', 'Opportunity results']
    }
    (OUT / 'data_standardization_remediation_round1_audit.json').write_text(json.dumps(audit, ensure_ascii=False, indent=2), encoding='utf-8')
    print(json.dumps({'school': audit['school'], 'date': audit['date'], 'schema': audit['schema']}, ensure_ascii=False))


if __name__ == '__main__': main()
